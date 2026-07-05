# analyze_training_progress.py
#
# Reads FRCNN training_metrics.xlsx and analyzes whether training has
# plateaued (enough) or is still improving (continue).
# FRCNN training_metrics.xlsx okur ve egitimin plato yapip yapmadigini
# (yeterli) ya da hala iyilesip iyilesmedigini (devam) analiz eder.
#
# Usage / Kullanim:
#     python analyze_training_progress.py

from pathlib import Path

import openpyxl


PROJECT_ROOT = Path("/home/atp-user-18/Desktop/uc_cihazlarda_terhmal_object_detection")
RUNS_NEW = PROJECT_ROOT / "runs_new"

MODELS = [
    "faster_rcnn_mobilenet_restratified_17.06.2026",
    "faster_rcnn_resnet50_restratified_17.06.20262",
]

# How many recent epochs without improvement counts as a plateau.
# Iyilesme olmadan kac son epoch plato sayilir.
PLATEAU_WINDOW = 10


def find_map_column(ws):
    # Find the column index holding overall mAP@0.5.
    # Genel mAP@0.5 tutan sutun indeksini bul.
    header = [cell.value for cell in ws[1]]
    for i, h in enumerate(header):
        if h and "map@0.5" in str(h).lower() and "0.95" not in str(h).lower():
            if "overall" in str(h).lower() or "(" not in str(h):
                return i, header
    # Fallback: any column with map@0.5.
    # Yedek: map@0.5 iceren herhangi bir sutun.
    for i, h in enumerate(header):
        if h and "map@0.5" in str(h).lower() and "0.95" not in str(h).lower():
            return i, header
    return None, header


def analyze_model(name):
    path = RUNS_NEW / name / "training_metrics.xlsx"
    if not path.exists():
        print(f"\n[{name}]")
        print(f"  training_metrics.xlsx bulunamadi")
        return

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    map_col, header = find_map_column(ws)
    if map_col is None:
        print(f"\n[{name}]")
        print(f"  mAP@0.5 sutunu bulunamadi. Header: {header}")
        return

    # Collect (epoch, map50) rows.
    # (epoch, map50) satirlarini topla.
    epochs = []
    maps = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            ep = int(row[0])
            mp = float(row[map_col])
        except (ValueError, TypeError):
            continue
        epochs.append(ep)
        maps.append(mp)

    if not maps:
        print(f"\n[{name}]")
        print(f"  Veri yok")
        return

    n = len(maps)
    best_map = max(maps)
    best_idx = maps.index(best_map)
    best_epoch = epochs[best_idx]
    last_epoch = epochs[-1]
    last_map = maps[-1]

    # Epochs since best.
    # Best'ten bu yana gecen epoch.
    epochs_since_best = last_epoch - best_epoch

    # Recent trend: best in last PLATEAU_WINDOW epochs.
    # Son trend: son PLATEAU_WINDOW epoch'ta en iyi.
    recent_maps = maps[-PLATEAU_WINDOW:] if n >= PLATEAU_WINDOW else maps
    recent_best = max(recent_maps)
    recent_improvement = recent_best - (recent_maps[0] if recent_maps else 0)

    print(f"\n{'=' * 60}")
    print(f"  {name}")
    print(f"{'=' * 60}")
    print(f"  Toplam epoch:        {n} (epoch {epochs[0]}-{last_epoch})")
    print(f"  Best mAP@0.5:        {best_map:.4f} (epoch {best_epoch})")
    print(f"  Son mAP@0.5:         {last_map:.4f} (epoch {last_epoch})")
    print(f"  Best'ten bu yana:    {epochs_since_best} epoch")
    print(f"  Son {len(recent_maps)} epoch iyilesme: {recent_improvement:+.4f}")

    # Last 10 epochs detail.
    # Son 10 epoch detayi.
    print(f"\n  Son {min(10, n)} epoch mAP@0.5:")
    for ep, mp in zip(epochs[-10:], maps[-10:]):
        marker = "  <- BEST" if mp == best_map else ""
        print(f"    epoch {ep:>3}: {mp:.4f}{marker}")

    # Verdict.
    # Karar.
    print(f"\n  DEGERLENDIRME:")
    if epochs_since_best >= PLATEAU_WINDOW:
        print(f"    PLATO: {epochs_since_best} epoch'tur iyilesme yok.")
        print(f"    Egitim YETERLI gorunuyor. Best epoch {best_epoch}'te yakalanmis.")
    elif recent_improvement > 0.01:
        print(f"    YUKSELISTE: son epoch'larda hala iyilesiyor (+{recent_improvement:.4f}).")
        print(f"    DEVAM etmeye deger.")
    else:
        print(f"    BELIRSIZ: net plato yok ama iyilesme de yavas.")
        print(f"    Birkac epoch daha denenebilir ama yakin.")


def main():
    print("FRCNN egitim ilerleme analizi")
    print(f"Plato penceresi: {PLATEAU_WINDOW} epoch")
    for name in MODELS:
        analyze_model(name)


if __name__ == "__main__":
    main()