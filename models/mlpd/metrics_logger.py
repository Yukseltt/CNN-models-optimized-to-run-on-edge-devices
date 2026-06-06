# metrics_logger.py
#
# Excel logger for MLPD-flavored RetinaNet training.
# MLPD-flavored RetinaNet egitimi icin Excel logger.
#
# MLPD's loss dict has two components: bbox_regression and classification.
# MLPD'nin loss dict'i iki bilesenli: bbox_regression ve classification.
#
# Layout / Yapi:
#     - Sheet 1 "Training Metrics" with one row per epoch.
#     - Sheet 1 "Training Metrics" her epoch icin bir satir.
#     - Sheet 2 "Confusion Matrix" placeholder for post-training analysis.
#     - Sheet 2 "Confusion Matrix" egitim sonrasi analiz icin placeholder.
#     - Sheet 3 "Charts" placeholder for charts.
#     - Sheet 3 "Charts" grafikler icin placeholder.

from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Static columns shared by all classes / Tum siniflarda ortak sabit sutunlar.

STATIC_HEADERS = [
    "Epoch", "Learning Rate",
    "Bbox Regression (Train)",
    "Classification (Train)",
    "Total Loss (Train)",
    "Bbox Regression (Val)",
    "Classification (Val)",
    "Total Loss (Val)",
    "Precision (Val)", "Recall (Val)", "F1 (Val)",
    "mAP@0.5 (overall)", "mAP@0.5:0.95 (overall)",
]


# Style constants / Stil sabitleri.

_THIN        = Side(border_style="thin", color="BBBBBB")
BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="EBF3FB")


# Helpers / Yardimcilar.

def _build_headers(class_names: list[str]) -> list[str]:
    # Build full header row by appending per-class columns.
    # Per-class sutunlari ekleyerek tam header satirini olustur.
    headers = list(STATIC_HEADERS)
    for name in class_names:
        headers.append(f"mAP@0.5 ({name})")
        headers.append(f"mAP@0.5:0.95 ({name})")
        headers.append(f"Precision ({name})")
        headers.append(f"Recall ({name})")
    return headers


def _apply_header(ws, class_names: list[str]) -> None:
    # Apply styled single-row header to the Training Metrics sheet.
    # Training Metrics sayfasina stilli tek satirli header uygular.
    headers = _build_headers(class_names)
    for ci, val in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for i in range(1, len(headers) + 1):
        width = 10 if i <= len(STATIC_HEADERS) else 18
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"


def _fmt(v) -> Optional[float]:
    # Round float values to 4 decimals, preserve None.
    # Float degerlerini 4 ondalikla yuvarlar, None'i korur.
    return None if v is None else round(float(v), 4)


def _sum_loss_components(losses: dict) -> Optional[float]:
    # Sum loss components, return None if dict is empty or all values None.
    # Loss component'leri toplar, dict bos veya tum degerler None ise None doner.
    parts = [
        losses.get("bbox_regression"),
        losses.get("classification"),
    ]
    valid = [v for v in parts if v is not None]
    if not valid:
        return None
    return sum(valid)


# Public API / Genel kullanim.

def build_excel(output_path: Path, class_names: list[str]) -> None:
    # Create a new Excel file with three sheets.
    # Uc sayfali yeni bir Excel dosyasi olusturur.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Training Metrics"
    _apply_header(ws, class_names)

    cm = wb.create_sheet("Confusion Matrix")
    cm["A1"] = "Confusion Matrix - Egitim tamamlandiktan sonra doldurulur."
    cm["A1"].font = Font(name="Arial", bold=True, size=11)

    ch = wb.create_sheet("Charts")
    for r, txt in enumerate([
        "Grafikler bu sayfaya eklenecektir.",
        "Insert -> Chart ile Training Metrics verilerini kullanabilirsiniz.",
    ], start=1):
        c = ch.cell(row=r, column=1, value=txt)
        c.font = Font(name="Arial", size=10, italic=True, color="555555")

    wb.save(output_path)


def append_epoch_row(
    output_path:       Path,
    epoch:             int,
    lr:                float,
    train_losses:      dict,
    val_losses:        dict,
    val_metrics:       dict,
    per_class_metrics: dict,
    class_names:       list[str],
) -> None:
    # Append a single epoch row to the existing Excel file.
    # Mevcut Excel dosyasina tek bir epoch satiri ekler.
    #
    # Args / Parametreler:
    #     train_losses: dict with keys bbox_regression, classification.
    #     train_losses: dict, anahtarlar bbox_regression, classification.
    #
    #     val_losses: same structure as train_losses, computed via hybrid pass.
    #     val_losses: train_losses ile ayni yapida, hybrid pass ile hesaplanir.
    #
    #     val_metrics: dict with keys precision, recall, f1, map50, map50_95.
    #     val_metrics: dict, anahtarlar precision, recall, f1, map50, map50_95.
    #
    #     per_class_metrics: dict class_name -> {map50, map50_95, precision, recall}.
    #     per_class_metrics: dict class_name -> {map50, map50_95, precision, recall}.
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Training Metrics"]
    data_row = ws.max_row + 1
    alt_fill = ALT_FILL if data_row % 2 == 0 else None

    train_total = _sum_loss_components(train_losses)
    val_total   = _sum_loss_components(val_losses)

    row = [
        epoch,
        round(lr, 8),
        _fmt(train_losses.get("bbox_regression")),
        _fmt(train_losses.get("classification")),
        _fmt(train_total),
        _fmt(val_losses.get("bbox_regression")),
        _fmt(val_losses.get("classification")),
        _fmt(val_total),
        _fmt(val_metrics.get("precision")),
        _fmt(val_metrics.get("recall")),
        _fmt(val_metrics.get("f1")),
        _fmt(val_metrics.get("map50")),
        _fmt(val_metrics.get("map50_95")),
    ]

    for name in class_names:
        m = per_class_metrics.get(name, {})
        row.append(_fmt(m.get("map50")))
        row.append(_fmt(m.get("map50_95")))
        row.append(_fmt(m.get("precision")))
        row.append(_fmt(m.get("recall")))

    for ci, val in enumerate(row, 1):
        c = ws.cell(row=data_row, column=ci, value=val)
        c.font = DATA_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill

    wb.save(output_path)
