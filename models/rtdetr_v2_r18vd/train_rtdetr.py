"""RT-DETR v2 fine-tune / resume training module.

Faster R-CNN modulu ile ayni run klasor yapisini uretir:

    runs/<NAME>/
        best.pt              # model state_dict + meta (best eval_map'e gore)
        last.pt              # son evaluation snapshot'i
        training_metrics.xlsx
        plots/
        checkpoints/         # HF Trainer checkpoint dirs (resume kaynagi)

Kullanim:
    from train_rtdetr import train
    train(data_dir=..., cfg={...})

RESUME_FROM: mevcut bir HF checkpoint dizini verilirse, Trainer oradan
optimizer/scheduler/RNG state'ini de yukleyerek devam eder. Yeni
checkpoint'ler runs/<NAME>/checkpoints/ altina yazilir.
"""

import sys
import time
from pathlib import Path
from typing import Optional

import json

import openpyxl
import torch
import torch.multiprocessing as mp

# Bu host'ta DataLoader worker'lari (persistent + prefetch) varsayilan
# "file_descriptor" paylasimi ile EMFILE (too many open files) veriyor.
# "file_system" stratejisi sart.
try:
    mp.set_sharing_strategy("file_system")
except RuntimeError:
    pass

from transformers import (
    AutoConfig,
    AutoImageProcessor,
    AutoModelForObjectDetection,
    EarlyStoppingCallback,
    Trainer,
    TrainerCallback,
    TrainerState,
    TrainingArguments,
)

_THIS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(_THIS_DIR))

from dataset import (
    LocalCocoDataset,
    ThermalCocoDataset,
    build_transforms,
    collate_fn,
)
from map_evaluator import MAPEvaluator
from metrics_logger import build_excel, append_row, build_row
from plot_rtdetr import make_plots


PROJECT_ROOT = _THIS_DIR.parent.parent
RUNS_DIR     = PROJECT_ROOT / "runs"


DEFAULT_CFG = {
    "CHECKPOINT":           "PekingU/rtdetr_v2_r18vd",
    "IMAGE_SIZE":           480,         # 640 -> 480: ~2.3x daha az piksel, ~2x throughput

    "EPOCHS":               150,
    "PER_DEVICE_TRAIN_BS":  16,
    "PER_DEVICE_EVAL_BS":   32,
    "GRAD_ACCUM":           1,

    "LR":                   5e-5,
    "MAX_GRAD_NORM":        0.1,
    "WARMUP_STEPS":         300,
    "WEIGHT_DECAY":         0.0,

    "WORKERS":              6,           # 12 -> 6: yeterli, fazla worker idle kaliyor
    "PREFETCH_FACTOR":      2,
    "EVAL_STEPS":           2000,        # ~her 1.5 epoch'ta bir eval (full epoch eval cok yavas)
    "SEED":                 42,
    "PATIENCE":             20,

    "PROJECT":              str(RUNS_DIR),
    "NAME":                 "rtdetr_v2_r18vd_default",
    "RESUME_FROM":          None,

    # True ise: HF Hub'dan SADECE config yuklenir, model agirliklari random
    # init edilir (hicbir pretrain yok). epoch 1'den kendi verinle baslar.
    # False ise: HF Hub'tan pretrain agirliklarla baslar (fine-tune).
    "FROM_SCRATCH":         False,

    # torch.compile: ~1.3-2x hizlanma; bazi modellerde uyumsuz olabilir.
    # Hata olursa otomatik kapanir, egitim fp normal seyrinde devam eder.
    "COMPILE":              True,
}


def _resolve_run_dir(project_dir: str, base_name: str) -> Path:
    project = Path(project_dir)
    project.mkdir(parents=True, exist_ok=True)
    if not (project / base_name).exists():
        return project / base_name
    counter = 2
    while (project / f"{base_name}{counter}").exists():
        counter += 1
    return project / f"{base_name}{counter}"


def _read_xlsx_last_progress(xlsx_path: Path) -> tuple[float, int]:
    """Mevcut xlsx'in son satirindan (epoch, step) doner; bos/yoksa (0.0, 0)."""
    if not xlsx_path.exists():
        return 0.0, 0
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb["Training Metrics"] if "Training Metrics" in wb.sheetnames else wb.active
        if ws.max_row < 2:
            wb.close()
            return 0.0, 0
        last = next(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))
        wb.close()
        ep   = float(last[0]) if last[0] is not None else 0.0
        step = int(last[1])   if last[1] is not None else 0
        return ep, step
    except Exception as e:
        print(f"[ExcelLog] son satir okunamadi ({e}); offset=0 ile devam.")
        return 0.0, 0


class _ExcelLogCallback(TrainerCallback):
    """Her evaluation'da training_metrics.xlsx'e bir satir yazar.

    Resume durumunda Trainer step/epoch counter'i 0'dan baslar; mevcut
    xlsx'in son satirindaki (epoch, step) offset olarak eklenir ki
    Excel'deki ilerleme kesintisiz devam etsin.
    """

    def __init__(
        self,
        xlsx_path:     Path,
        class_names:   list[str],
        epoch_offset:  float = 0.0,
        step_offset:   int   = 0,
    ):
        self.xlsx_path     = Path(xlsx_path)
        self.class_names   = class_names
        self.epoch_offset  = float(epoch_offset)
        self.step_offset   = int(step_offset)
        self._last_train_loss = None
        self._last_lr = None
        if not self.xlsx_path.exists():
            build_excel(self.xlsx_path, class_names)
        print(f"[ExcelLog] -> {self.xlsx_path}")
        if self.epoch_offset or self.step_offset:
            print(
                f"[ExcelLog] resume offset: epoch+={self.epoch_offset:.4f}, "
                f"step+={self.step_offset}"
            )

    def on_log(self, args, state, control, logs=None, **kwargs):
        if not logs:
            return
        if "loss" in logs and "eval_loss" not in logs:
            self._last_train_loss = logs["loss"]
        if "learning_rate" in logs:
            self._last_lr = logs["learning_rate"]

    def on_evaluate(self, args, state, control, metrics=None, **kwargs):
        if metrics is None:
            return
        cur_epoch = (state.epoch or 0.0) + self.epoch_offset
        cur_step  = state.global_step + self.step_offset
        row = build_row(
            epoch=cur_epoch,
            step=cur_step,
            lr=self._last_lr,
            train_loss=self._last_train_loss,
            eval_metrics=metrics,
            class_names=self.class_names,
        )
        # Excel yazimi HICBIR kosulda egitimi olduremez: Drive FUSE mount
        # gecici olarak dosyayi kaybedebilir (30 saatlik run'i 1 satir log
        # icin cokertmek kabul edilemez). append_row zaten retry + yeniden
        # kurma yapiyor; o da yetmezse satiri atla, egitime devam.
        try:
            append_row(self.xlsx_path, row, class_names=self.class_names)
        except Exception as e:
            print(f"[ExcelLog] UYARI: satir yazilamadi ({type(e).__name__}: {e}); atlandi.")
        print(
            f"[ExcelLog] epoch={row[0]} step={row[1]} "
            f"train_loss={row[3]} val_loss={row[4]} "
            f"mAP={row[5]} mAP@0.5={row[6]}"
        )


class _StepTimingCallback(TrainerCallback):
    """Her N step'te bir gercek hizi (step/s, sample/s, ETA) basar."""

    def __init__(self, batch_size: int, log_every: int = 50):
        self.batch_size = batch_size
        self.log_every  = log_every
        self._last_t    = None
        self._last_step = 0

    def on_step_end(self, args, state, control, **kwargs):
        import time
        if state.global_step % self.log_every != 0:
            return
        now = time.time()
        if self._last_t is not None:
            dt    = now - self._last_t
            steps = state.global_step - self._last_step
            if steps > 0 and dt > 0:
                sps   = steps / dt
                samps = sps * self.batch_size
                remain_steps = state.max_steps - state.global_step
                eta_s = remain_steps / sps if sps > 0 else 0
                eta_h = eta_s / 3600
                print(
                    f"[Timing] step={state.global_step}/{state.max_steps} "
                    f"step/s={sps:.2f} sample/s={samps:.1f} ETA={eta_h:.2f}h"
                )
        self._last_t    = now
        self._last_step = state.global_step


class _BestLastPtCallback(TrainerCallback):
    """Her evaluation sonrasi last.pt ve (iyilesirse) best.pt yazar.

    Faster R-CNN ile uyumlu sema: {epoch, model, cfg, class_names, metric}.
    """

    def __init__(
        self,
        run_dir:      Path,
        cfg:          dict,
        class_names:  list[str],
        metric_key:   str = "eval_map",
        epoch_offset: float = 0.0,
    ):
        self.run_dir      = Path(run_dir)
        self.cfg          = cfg
        self.class_names  = class_names
        self.metric_key   = metric_key
        self.epoch_offset = float(epoch_offset)
        self._best_metric = -1.0

    def _snapshot(self, model, epoch, metric):
        return {
            "epoch":       int(epoch) if epoch is not None else -1,
            "model":       model.state_dict(),
            "cfg":         self.cfg,
            "class_names": self.class_names,
            "metric":      float(metric) if metric is not None else None,
        }

    @staticmethod
    def _safe_save(obj, path: Path) -> bool:
        """torch.save — Drive FUSE gecici hatalarina karsi 3 deneme.
        Basarisizsa False doner; egitim devam eder (snapshot bir sonraki
        eval'de zaten yeniden yazilir)."""
        last_err = None
        for delay in (0, 5, 15):
            if delay:
                time.sleep(delay)
            try:
                torch.save(obj, path)
                return True
            except Exception as e:
                last_err = e
        print(f"[BestLast] UYARI: {path.name} yazilamadi "
              f"({type(last_err).__name__}: {last_err}); atlandi.")
        return False

    def on_evaluate(self, args, state, control, metrics=None, model=None, **kwargs):
        if model is None or metrics is None:
            return
        metric = metrics.get(self.metric_key)
        epoch  = (state.epoch or 0.0) + self.epoch_offset

        last_path = self.run_dir / "last.pt"
        self._safe_save(self._snapshot(model, epoch, metric), last_path)

        if metric is not None and metric > self._best_metric:
            best_path = self.run_dir / "best.pt"
            # _best_metric sadece yazim BASARILIYSA guncellenir; yoksa diskteki
            # best.pt eski kalir ve bir sonraki iyilesmede tekrar denenir.
            if self._safe_save(self._snapshot(model, epoch, metric), best_path):
                self._best_metric = float(metric)
                print(f"[BestLast] NEW BEST {self.metric_key}={metric:.4f} -> {best_path.name}")


class _ResumeStateTrainer(Trainer):
    """Trainer alt sinifi: resume'da sadece trainer_state.json'dan
    global_step / epoch'u alir; model'in kendisini ve optimizer/scheduler/RNG'yi
    yuklemez (model dis taraftan from_pretrained ile dogru sekilde yuklenmistir).

    Boylece eski transformers semasiyla kaydedilmis checkpoint'lerin
    state_dict key mismatch sorunundan etkilenmeden, Trainer progress
    counter'i dogru epoch/step'ten devam eder.
    """

    def __init__(self, *args, resume_state_path: Optional[str] = None, **kwargs):
        super().__init__(*args, **kwargs)
        self._resume_state_path = resume_state_path

    def _load_from_checkpoint(self, resume_from_checkpoint, model=None):
        # Model agirliklarini yeniden yukleme — disarida from_pretrained
        # zaten dogru key conversion'la yukledi.
        return

    def _load_optimizer_and_scheduler(self, checkpoint):
        # Optimizer/scheduler/RNG yuklemesini atla — fresh basla.
        return

    def _load_rng_state(self, checkpoint):
        return


def _read_trainer_state(state_json_path: Path) -> tuple[float, int]:
    """trainer_state.json'dan (epoch, global_step) doner."""
    try:
        with open(state_json_path) as f:
            st = json.load(f)
        return float(st.get("epoch", 0.0) or 0.0), int(st.get("global_step", 0) or 0)
    except Exception as e:
        print(f"[Resume] trainer_state.json okunamadi ({e}); 0'dan baslanir.")
        return 0.0, 0


def _load_initial_best(run_dir: Path, metric_key: str = "eval_map") -> float:
    """Resume oncesi best.pt varsa metric'i oku, yoksa -1.0."""
    best = run_dir / "best.pt"
    if not best.exists():
        return -1.0
    try:
        ckpt = torch.load(best, map_location="cpu", weights_only=False)
        m = ckpt.get("metric")
        return float(m) if m is not None else -1.0
    except Exception as e:
        print(f"[BestLast] best.pt okunamadi ({e}), -1.0 olarak basliyor.")
        return -1.0


def train(data_dir: str, cfg: Optional[dict] = None) -> None:
    c = {**DEFAULT_CFG, **(cfg or {})}

    torch.manual_seed(c["SEED"])
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(c["SEED"])

    is_resume       = c.get("RESUME_FROM") is not None
    is_from_scratch = bool(c.get("FROM_SCRATCH"))

    if is_resume and is_from_scratch:
        raise ValueError("RESUME_FROM ve FROM_SCRATCH ayni anda kullanilamaz.")

    print(f"[RT-DETR] Checkpoint: {c['CHECKPOINT']}")
    print(f"[RT-DETR] Image size: {c['IMAGE_SIZE']}")
    print(f"[RT-DETR] Epochs:     {c['EPOCHS']}, batch={c['PER_DEVICE_TRAIN_BS']}, "
          f"grad_accum={c['GRAD_ACCUM']}")
    print(f"[RT-DETR] LR:         {c['LR']}, warmup={c['WARMUP_STEPS']}")
    if is_resume:
        print(f"[RT-DETR] Resume:     {c['RESUME_FROM']}")

    # Egitime baslamadan once kullaniciya mod sorulur (resume haricinde).
    # 1 -> pretrain agirliklarla fine-tune    (FROM_SCRATCH=False)
    # 0 -> sifirdan, random init              (FROM_SCRATCH=True)
    # Background runlar icin SKIP_CONFIRM=True ise config'teki FROM_SCRATCH kullanilir.
    if not is_resume and not c.get("SKIP_CONFIRM", False):
        print()
        print("=" * 70)
        print("Egitim modu sec:")
        print("  1 -> Pretrain  (HF Hub'tan pretrained agirliklarla fine-tune)")
        print("  0 -> Sifirdan  (random init, hicbir pretrain yok)")
        print("=" * 70)
        while True:
            try:
                ans = input("Secim [1/0]: ").strip()
            except EOFError:
                ans = ""
            if ans == "1":
                is_from_scratch = False
                break
            if ans == "0":
                is_from_scratch = True
                break
            print(f"Gecersiz: {ans!r}. Lutfen 1 veya 0 gir (iptal icin Ctrl+C).")
        c["FROM_SCRATCH"] = is_from_scratch

    if is_from_scratch:
        print("[RT-DETR] Mod: SIFIRDAN (random init, pretrain YOK)")
    elif not is_resume:
        print("[RT-DETR] Mod: PRETRAIN fine-tune")

    # Resume modunda: mevcut run_dir'i yeniden kullan
    # (RESUME_FROM = .../runs/<NAME>/checkpoints/checkpoint-N varsayimi)
    # Boylece Excel append edilir, best/last.pt ayni dizinde guncel kalir.
    if is_resume:
        resume_path = Path(c["RESUME_FROM"]).resolve()
        if resume_path.name.startswith("checkpoint-") and resume_path.parent.name == "checkpoints":
            run_dir = resume_path.parent.parent
            hf_output_dir = resume_path.parent
            print(f"[RT-DETR] Resume run_dir tespit edildi: {run_dir}")
        else:
            run_dir = _resolve_run_dir(c["PROJECT"], c["NAME"])
            hf_output_dir = run_dir / "checkpoints"
        run_dir.mkdir(parents=True, exist_ok=True)
        hf_output_dir.mkdir(parents=True, exist_ok=True)
    else:
        run_dir = _resolve_run_dir(c["PROJECT"], c["NAME"])
        run_dir.mkdir(parents=True, exist_ok=True)
        hf_output_dir = run_dir / "checkpoints"
        hf_output_dir.mkdir(parents=True, exist_ok=True)
    print(f"[RT-DETR] Run dir:    {run_dir}")

    train_aug, val_aug = build_transforms()

    train_base = LocalCocoDataset(Path(data_dir) / "train")
    val_base   = LocalCocoDataset(Path(data_dir) / "val")
    class_names = train_base.category_names
    id2label    = {i: n for i, n in enumerate(class_names)}
    label2id    = {n: i for i, n in id2label.items()}
    print(f"[RT-DETR] Train: {len(train_base)} imgs, Val: {len(val_base)} imgs")
    print(f"[RT-DETR] Classes: {class_names}")

    # Image processor: her zaman HF Hub'tan / checkpoint'ten yuklenir.
    # Bunlar agirlik degil; sadece preprocessing (resize, normalize) parametreleri.
    proc_source = c["RESUME_FROM"] if is_resume else c["CHECKPOINT"]
    image_processor = AutoImageProcessor.from_pretrained(
        proc_source,
        do_resize=True,
        size={"width": c["IMAGE_SIZE"], "height": c["IMAGE_SIZE"]},
        use_fast=True,
    )

    if is_resume:
        # Resume: checkpoint'ten model agirliklarini yukle.
        model = AutoModelForObjectDetection.from_pretrained(c["RESUME_FROM"])
    elif is_from_scratch:
        # From scratch: sadece config yukle, agirliklari RANDOM init et.
        config = AutoConfig.from_pretrained(c["CHECKPOINT"])
        config.id2label   = id2label
        config.label2id   = label2id
        config.num_labels = len(id2label)
        model = AutoModelForObjectDetection.from_config(config)
        n_params = sum(p.numel() for p in model.parameters())
        print(f"[RT-DETR] Random init model: {n_params/1e6:.2f}M params")
    else:
        # Fine-tune: HF Hub'tan pretrained agirliklarla basla.
        model = AutoModelForObjectDetection.from_pretrained(
            c["CHECKPOINT"],
            id2label=id2label,
            label2id=label2id,
            ignore_mismatched_sizes=True,
        )

    # torch.compile: PyTorch 2.x JIT compile.
    # NOT: RT-DETR v2'nin encoder'i "reduce-overhead" modu ile uyumsuz
    # (encoder forward'inda dynamic control flow var, CUDA graph capture'da
    # NoneType unpack hatasi). "default" modu daha guvenli ama daha az kazanc.
    # Sorun yasarsan COMPILE=False yap.
    compile_mode = c.get("COMPILE_MODE", "default")  # "default" | "reduce-overhead" | "max-autotune"
    if c.get("COMPILE") and torch.cuda.is_available():
        try:
            model = torch.compile(model, mode=compile_mode, dynamic=True)
            print(f"[RT-DETR] torch.compile aktif (mode={compile_mode}, dynamic=True)")
        except Exception as e:
            print(f"[RT-DETR] torch.compile basarisiz: {e}. Devre disi.")

    train_ds = ThermalCocoDataset(train_base, image_processor, transform=train_aug)
    val_ds   = ThermalCocoDataset(val_base,   image_processor, transform=val_aug)

    use_bf16 = torch.cuda.is_available() and torch.cuda.is_bf16_supported()
    use_fp16 = torch.cuda.is_available() and not use_bf16
    # TF32 sadece Ampere+ (compute capability >= 8.0) GPU'larda mevcut (T4 = 7.5).
    use_tf32 = torch.cuda.is_available() and torch.cuda.get_device_capability()[0] >= 8

    # TF32 fp32 matmul'lerini hizlandirir (H200 / Ampere+). bf16 patikalarini
    # etkilemez ama optimizer / norm gibi fp32 kalan kisimlarda kazanc saglar.
    if torch.cuda.is_available():
        torch.set_float32_matmul_precision("high")
        torch.backends.cudnn.benchmark = True

    training_args = TrainingArguments(
        output_dir                  = str(hf_output_dir),
        num_train_epochs            = c["EPOCHS"],
        max_grad_norm               = c["MAX_GRAD_NORM"],
        learning_rate               = c["LR"],
        warmup_steps                = c["WARMUP_STEPS"],
        weight_decay                = c["WEIGHT_DECAY"],

        per_device_train_batch_size = c["PER_DEVICE_TRAIN_BS"],
        per_device_eval_batch_size  = c["PER_DEVICE_EVAL_BS"],
        gradient_accumulation_steps = c["GRAD_ACCUM"],

        dataloader_num_workers      = c["WORKERS"],
        dataloader_pin_memory       = True,
        dataloader_persistent_workers = c["WORKERS"] > 0,
        dataloader_prefetch_factor  = c.get("PREFETCH_FACTOR", 4) if c["WORKERS"] > 0 else None,
        dataloader_drop_last        = True,

        bf16                        = use_bf16,
        fp16                        = use_fp16,
        bf16_full_eval              = use_bf16,
        tf32                        = use_tf32,

        # Fused AdamW: tek CUDA kernel, ~5-10% adim hizlandirma.
        optim                       = "adamw_torch_fused",

        # Eval'de pin'lenmis tensor'leri parça parça topla -> bellek piki azal.
        eval_accumulation_steps     = c.get("EVAL_ACCUM_STEPS", 4),

        metric_for_best_model       = "eval_map",
        greater_is_better           = True,
        load_best_model_at_end      = False,
        eval_strategy               = "steps",
        eval_steps                  = c.get("EVAL_STEPS", 2000),
        save_strategy               = "steps",
        save_steps                  = c.get("EVAL_STEPS", 2000),
        save_total_limit            = 2,
        logging_steps               = 50,
        logging_first_step          = True,
        remove_unused_columns       = False,
        eval_do_concat_batches      = False,
        report_to                   = "none",
        push_to_hub                 = False,
        seed                        = c["SEED"],
    )

    # Resume mantigi:
    #   - Model agirliklari: from_pretrained(RESUME_FROM) ile dogru key
    #     conversion uygulanarak zaten yuklendi (yukarida).
    #   - trainer_state.json'dan global_step/epoch okunur ve Trainer state'ine
    #     enjekte edilir; boylece progress bar/eval_steps tetikleyici tutarli olur.
    #   - Optimizer/scheduler/RNG fresh — bilinen tradeoff.
    # state.global_step zaten dogru baslayacagi icin Excel offset'i 0.
    epoch_offset, step_offset = 0.0, 0
    resume_state_path = None
    if is_resume:
        st_json = Path(c["RESUME_FROM"]) / "trainer_state.json"
        if st_json.exists():
            resume_state_path = str(st_json)
            ep, gs = _read_trainer_state(st_json)
            print(f"[Resume] trainer_state.json: epoch={ep:.4f} global_step={gs}")
        else:
            print(f"[Resume] trainer_state.json yok ({st_json}); Excel'in son satirindan offset alinacak.")
            epoch_offset, step_offset = _read_xlsx_last_progress(
                run_dir / "training_metrics.xlsx"
            )

    excel_cb = _ExcelLogCallback(
        run_dir / "training_metrics.xlsx", class_names,
        epoch_offset=epoch_offset, step_offset=step_offset,
    )
    bestlast_cb = _BestLastPtCallback(
        run_dir, c, class_names, epoch_offset=epoch_offset,
    )
    bestlast_cb._best_metric = _load_initial_best(run_dir)
    early_cb = EarlyStoppingCallback(early_stopping_patience=c["PATIENCE"])

    eval_metrics_fn = MAPEvaluator(
        image_processor=image_processor, threshold=0.0, id2label=id2label,
    )

    trainer_cls = _ResumeStateTrainer if is_resume else Trainer
    trainer_kwargs = dict(
        model=model,
        args=training_args,
        train_dataset=train_ds,
        eval_dataset=val_ds,
        processing_class=image_processor,
        data_collator=collate_fn,
        compute_metrics=eval_metrics_fn,
        callbacks=[
            excel_cb,
            bestlast_cb,
            early_cb,
            _StepTimingCallback(
                batch_size=c["PER_DEVICE_TRAIN_BS"] * c["GRAD_ACCUM"],
                log_every=50,
            ),
        ],
    )
    if is_resume:
        trainer_kwargs["resume_state_path"] = resume_state_path
    trainer = trainer_cls(**trainer_kwargs)

    # Resume: trainer.train(resume_from_checkpoint=path) cagrilir, ama
    # _ResumeStateTrainer subclass'i model/optimizer/scheduler/RNG yuklemesini
    # atlar. HF Trainer trainer_state.json'dan state'i yine de okur — bu
    # sayede state.global_step ve state.epoch dogru baslar, progress bar
    # dogru ilerler, eval_steps tetikleyici tutarli kalir.
    # Model agirliklari zaten from_pretrained ile yuklendi.
    if is_resume:
        print(
            f"[RT-DETR] Egitim basliyor (state resume from {c['RESUME_FROM']}, "
            "optimizer/scheduler fresh)...\n"
        )
        trainer.train(resume_from_checkpoint=c["RESUME_FROM"])
    else:
        print(f"[RT-DETR] Egitim basliyor (fresh)...\n")
        trainer.train()

    print("\n[RT-DETR] Egitim tamamlandi. Grafikler uretiliyor...")
    try:
        make_plots(run_dir)
    except Exception as e:
        print(f"[RT-DETR] Grafik uretiminde hata: {e}")

    print(f"\n[RT-DETR] Bitti. Run dizini: {run_dir}")
