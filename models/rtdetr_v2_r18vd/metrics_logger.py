"""Excel logger for RT-DETR v2 training.

Bir epoch icin tek satir yazar; Faster R-CNN run klasoru ile ayni stiilde
(training_metrics.xlsx) cikti uretir, fakat RT-DETR'nin metric setine gore
sutunlar farklidir (torchmetrics MeanAveragePrecision ciktilari).
"""

from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


STATIC_HEADERS = [
    "Epoch", "Step", "Learning Rate",
    "Train Loss", "Val Loss",
    "mAP@0.5:0.95", "mAP@0.5", "mAP@0.75",
    "mAR@1", "mAR@10", "mAR@100",
]

_THIN       = Side(border_style="thin", color="BBBBBB")
BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="EBF3FB")


def _build_headers(class_names: list[str]) -> list[str]:
    h = list(STATIC_HEADERS)
    for n in class_names:
        h.append(f"mAP ({n})")
        h.append(f"mAR@100 ({n})")
    return h


def build_excel(path: Path, class_names: list[str]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Training Metrics"
    headers = _build_headers(class_names)
    for ci, val in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = (
            12 if i <= len(STATIC_HEADERS) else 18
        )
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"
    wb.create_sheet("Charts")
    wb.save(path)


def _fmt(v) -> Optional[float]:
    return None if v is None else round(float(v), 6)


def append_row(path: Path, row: list) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb["Training Metrics"]
    r = ws.max_row + 1
    alt = ALT_FILL if r % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = DATA_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        if alt:
            c.fill = alt
    wb.save(path)


def build_row(
    epoch:           Optional[float],
    step:            int,
    lr:              Optional[float],
    train_loss:      Optional[float],
    eval_metrics:    dict,
    class_names:     list[str],
) -> list:
    m = eval_metrics or {}
    row = [
        round(epoch, 4) if epoch is not None else None,
        step,
        _fmt(lr),
        _fmt(train_loss),
        _fmt(m.get("eval_loss")),
        _fmt(m.get("eval_map")),
        _fmt(m.get("eval_map_50")),
        _fmt(m.get("eval_map_75")),
        _fmt(m.get("eval_mar_1")),
        _fmt(m.get("eval_mar_10")),
        _fmt(m.get("eval_mar_100")),
    ]
    for n in class_names:
        row.append(_fmt(m.get(f"eval_map_{n}")))
        row.append(_fmt(m.get(f"eval_mar_100_{n}")))
    return row
