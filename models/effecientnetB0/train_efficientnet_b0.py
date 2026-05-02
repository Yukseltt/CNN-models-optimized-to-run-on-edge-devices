# train_efficientnet_b0.py
#
# Training script for EfficientNet-B0 on COCO bbox crops (thermal classifier).
# COCO bbox kirpimlari uzerinde EfficientNet-B0 egitim scripti (termal siniflandirici).
#
# Same training-loop conventions as faster_rcnn_dla34/train_faster_rcnn_dla34.py:
# manual warmup + cosine LR, per-epoch validation, Excel logging, best/last
# checkpointing, early stopping, resume from last.pt.
# faster_rcnn_dla34/train_faster_rcnn_dla34.py ile ayni egitim kuralllari:
# manuel warmup + cosine LR, her epoch validation, Excel logging, best/last
# checkpoint, early stopping, last.pt'den resume.

import json
import math
import random
import time
from itertools import chain
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from PIL.Image import BICUBIC
from torch.amp import GradScaler, autocast
from torch.optim import SGD
from torch.utils.data import DataLoader, Dataset
from torchvision.transforms import (
    Compose,
    Normalize,
    RandomCrop,
    RandomHorizontalFlip,
    Resize,
    ToTensor,
)

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))
from model_factory import create_efficientnet_b0, count_parameters


# Path constants / Yol sabitleri.

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
RUNS_DIR     = PROJECT_ROOT / "runs"


# Default config / Varsayilan konfigurasyon.

DEFAULT_CFG = {
    "EPOCHS":         50,
    "BATCH_SIZE":     64,
    "PATIENCE":       10,
    "WORKERS":        4,
    "SEED":           0,
    "DEVICE":         "cuda:0",

    # Tek bir base LR; param gruplari (stem+blocks, head_pre, head_cls)
    # icin notebook'taki carpanlar kullanilir: 0.1 / 0.2 / 1.0.
    "LR0":            0.01,
    "LRF":            0.01,
    "MOMENTUM":       0.9,
    "WEIGHT_DECAY":   0.001,
    "NESTEROV":       True,
    "WARMUP_EPOCHS":  2,

    "BACKBONE_LR_MULT": 0.1,
    "HEAD_PRE_LR_MULT": 0.2,
    "HEAD_CLS_LR_MULT": 1.0,

    "PRETRAINED":      True,
    "PRETRAINED_PATH": None,
    "INPUT_SIZE":      224,
    "MIN_BBOX_SIDE":   4,
    "USE_AMP":         True,

    "PROJECT":     str(RUNS_DIR),
    "NAME":        "efficientnet_b0_default",
    "OUTPUT_XLSX": "training_metrics.xlsx",

    "LOG_INTERVAL": 50,

    # Bir checkpoint yoluna ayarlanirsa, oradan devam eder.
    "RESUME_FROM": None,
}


# Helpers / Yardimcilar.

def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def compute_lr(
    epoch:         int,
    total_epochs:  int,
    warmup_epochs: int,
    lr0:           float,
    lrf:           float,
) -> float:
    if epoch < warmup_epochs:
        return lr0 * (epoch + 1) / max(1, warmup_epochs)
    progress = (epoch - warmup_epochs) / max(1, total_epochs - warmup_epochs)
    cosine   = 0.5 * (1 + math.cos(math.pi * progress))
    return lr0 * (lrf + (1 - lrf) * cosine)


def resolve_run_dir(project_dir: str, base_name: str) -> Path:
    project = Path(project_dir)
    project.mkdir(parents=True, exist_ok=True)

    if not (project / base_name).exists():
        return project / base_name

    counter = 2
    while (project / f"{base_name}{counter}").exists():
        counter += 1
    return project / f"{base_name}{counter}"


# Dataset / Dataset.

class CocoBBoxClassifier(Dataset):
    # Her COCO annotation bir bbox kirpimi olarak okunur ve siniflandirma
    # ornegi olur. Dataset notebook'takiyle ayni mantikta.

    def __init__(
        self,
        ann_path:   str,
        images_dir: str,
        transform=None,
        min_side:   int = 4,
    ):
        with open(ann_path, "r") as f:
            data = json.load(f)

        self.images_dir = Path(images_dir)
        self.transform  = transform

        cats = sorted(data["categories"], key=lambda c: c["id"])
        self.cat_id_to_idx = {c["id"]: i for i, c in enumerate(cats)}
        self.class_names   = [c["name"] for c in cats]

        id_to_file = {im["id"]: im["file_name"] for im in data["images"]}

        self.samples = []
        for ann in data["annotations"]:
            x, y, w, h = ann["bbox"]
            if w < min_side or h < min_side:
                continue
            self.samples.append((
                id_to_file[ann["image_id"]],
                (float(x), float(y), float(w), float(h)),
                self.cat_id_to_idx[ann["category_id"]],
            ))

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, idx: int):
        fname, (x, y, w, h), cls = self.samples[idx]
        img  = Image.open(self.images_dir / fname).convert("RGB")
        crop = img.crop((x, y, x + w, y + h))
        if self.transform is not None:
            crop = self.transform(crop)
        return crop, cls


def build_train_transform(input_size: int):
    return Compose([
        Resize((input_size + 32, input_size + 32), BICUBIC),
        RandomCrop(input_size),
        RandomHorizontalFlip(),
        ToTensor(),
        Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
    ])


def build_eval_transform(input_size: int):
    return Compose([
        Resize((input_size, input_size), BICUBIC),
        ToTensor(),
        Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
    ])


def build_dataloaders(data_dir: str, cfg: dict) -> Tuple[DataLoader, DataLoader, list]:
    data_dir = Path(data_dir)

    train_ds = CocoBBoxClassifier(
        ann_path=str(data_dir / "train" / "_annotations.coco.json"),
        images_dir=str(data_dir / "train" / "images"),
        transform=build_train_transform(cfg["INPUT_SIZE"]),
        min_side=cfg["MIN_BBOX_SIDE"],
    )
    val_ds = CocoBBoxClassifier(
        ann_path=str(data_dir / "val" / "_annotations.coco.json"),
        images_dir=str(data_dir / "val" / "images"),
        transform=build_eval_transform(cfg["INPUT_SIZE"]),
        min_side=cfg["MIN_BBOX_SIDE"],
    )

    train_loader = DataLoader(
        train_ds,
        batch_size=cfg["BATCH_SIZE"],
        shuffle=True,
        num_workers=cfg["WORKERS"],
        pin_memory=True,
        drop_last=True,
        persistent_workers=(cfg["WORKERS"] > 0),
    )
    val_loader = DataLoader(
        val_ds,
        batch_size=cfg["BATCH_SIZE"],
        shuffle=False,
        num_workers=cfg["WORKERS"],
        pin_memory=True,
        drop_last=False,
        persistent_workers=(cfg["WORKERS"] > 0),
    )
    return train_loader, val_loader, train_ds.class_names


# Optimizer / Optimizer.

def build_optimizer(model: nn.Module, cfg: dict) -> torch.optim.Optimizer:
    # 3 param grubu: stem+blocks (backbone), head[:6] (head_pre), head[6] (head_cls).
    # Notebook'tan: backbone *0.1, head_pre *0.2, head_cls *1.0.
    return SGD(
        [
            {
                "params": list(chain(model.stem.parameters(), model.blocks.parameters())),
                "lr":     cfg["LR0"] * cfg["BACKBONE_LR_MULT"],
                "name":   "backbone",
            },
            {
                "params": list(model.head[:6].parameters()),
                "lr":     cfg["LR0"] * cfg["HEAD_PRE_LR_MULT"],
                "name":   "head_pre",
            },
            {
                "params": list(model.head[6].parameters()),
                "lr":     cfg["LR0"] * cfg["HEAD_CLS_LR_MULT"],
                "name":   "head_cls",
            },
        ],
        momentum=cfg["MOMENTUM"],
        weight_decay=cfg["WEIGHT_DECAY"],
        nesterov=cfg["NESTEROV"],
    )


def set_param_group_lrs(optimizer: torch.optim.Optimizer, base_lr: float, cfg: dict) -> None:
    mults = {
        "backbone": cfg["BACKBONE_LR_MULT"],
        "head_pre": cfg["HEAD_PRE_LR_MULT"],
        "head_cls": cfg["HEAD_CLS_LR_MULT"],
    }
    for g in optimizer.param_groups:
        g["lr"] = base_lr * mults[g["name"]]


# Excel logger / Excel logger.

HEADERS = [
    "Epoch", "Learning Rate",
    "Train Loss", "Train Accuracy", "Train Precision", "Train Recall",
    "Val Loss",   "Val Accuracy",   "Val Precision",   "Val Recall",
]

_THIN       = Side(border_style="thin", color="BBBBBB")
BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)


def build_excel(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Training Metrics"
    for ci, val in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"
    wb.save(path)


def append_epoch_row(path: Path, values: list) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb["Training Metrics"]
    r  = ws.max_row + 1
    alt = PatternFill("solid", start_color="EBF3FB") if r % 2 == 0 else None
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font      = DATA_FONT
        c.alignment = Alignment(horizontal="center")
        c.border    = BORDER
        if alt:
            c.fill = alt
    wb.save(path)


def _fmt(v):
    return None if v is None else round(float(v), 4)


# Metrics / Metrikler.

def compute_classification_metrics(
    preds:       torch.Tensor,
    targets:     torch.Tensor,
    num_classes: int,
) -> dict:
    # Macro-averaged precision/recall + accuracy.
    # Macro ortalamali precision/recall + accuracy.
    correct  = (preds == targets).float().sum().item()
    total    = targets.numel()
    accuracy = correct / total if total > 0 else 0.0

    precisions = []
    recalls    = []
    for c in range(num_classes):
        tp = ((preds == c) & (targets == c)).float().sum().item()
        fp = ((preds == c) & (targets != c)).float().sum().item()
        fn = ((preds != c) & (targets == c)).float().sum().item()
        p  = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        r  = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        precisions.append(p)
        recalls.append(r)

    return {
        "accuracy":  accuracy,
        "precision": float(np.mean(precisions)),
        "recall":    float(np.mean(recalls)),
    }


# Train epoch / Train epoch.

def train_one_epoch(
    model:        nn.Module,
    optimizer:    torch.optim.Optimizer,
    criterion:    nn.Module,
    loader:       DataLoader,
    device:       str,
    epoch:        int,
    scaler:       Optional[GradScaler],
    use_amp:      bool,
    log_interval: int,
    num_classes:  int,
) -> dict:
    model.train()

    loss_sum  = 0.0
    n_samples = 0

    all_preds:   list = []
    all_targets: list = []

    epoch_start = time.time()

    for batch_idx, (images, labels) in enumerate(loader):
        images = images.to(device, non_blocking=True)
        labels = labels.to(device, non_blocking=True)

        optimizer.zero_grad(set_to_none=True)

        if use_amp:
            with autocast("cuda"):
                logits = model(images)
                loss   = criterion(logits, labels)
            scaler.scale(loss).backward()
            scaler.step(optimizer)
            scaler.update()
        else:
            logits = model(images)
            loss   = criterion(logits, labels)
            loss.backward()
            optimizer.step()

        with torch.no_grad():
            preds = logits.argmax(dim=1)

        bs        = images.size(0)
        loss_sum += loss.item() * bs
        n_samples += bs
        all_preds.append(preds.detach().cpu())
        all_targets.append(labels.detach().cpu())

        if (batch_idx + 1) % log_interval == 0:
            elapsed  = time.time() - epoch_start
            avg_loss = loss_sum / max(1, n_samples)
            print(
                f"  [Epoch {epoch}] batch {batch_idx + 1}/{len(loader)}  "
                f"avg_loss={avg_loss:.4f}  elapsed={elapsed:.1f}s"
            )

    all_preds_t   = torch.cat(all_preds)
    all_targets_t = torch.cat(all_targets)
    metrics       = compute_classification_metrics(all_preds_t, all_targets_t, num_classes)

    return {
        "loss":      loss_sum / max(1, n_samples),
        "accuracy":  metrics["accuracy"],
        "precision": metrics["precision"],
        "recall":    metrics["recall"],
    }


# Validation / Validation.

@torch.no_grad()
def validate(
    model:       nn.Module,
    criterion:   nn.Module,
    loader:      DataLoader,
    device:      str,
    num_classes: int,
) -> dict:
    model.eval()

    loss_sum  = 0.0
    n_samples = 0

    all_preds:   list = []
    all_targets: list = []

    val_start = time.time()

    for images, labels in loader:
        images = images.to(device, non_blocking=True)
        labels = labels.to(device, non_blocking=True)

        logits = model(images)
        loss   = criterion(logits, labels)
        preds  = logits.argmax(dim=1)

        bs         = images.size(0)
        loss_sum  += loss.item() * bs
        n_samples += bs
        all_preds.append(preds.cpu())
        all_targets.append(labels.cpu())

    all_preds_t   = torch.cat(all_preds)
    all_targets_t = torch.cat(all_targets)
    metrics       = compute_classification_metrics(all_preds_t, all_targets_t, num_classes)

    elapsed = time.time() - val_start
    print(f"  Validation done in {elapsed:.1f}s")

    return {
        "loss":      loss_sum / max(1, n_samples),
        "accuracy":  metrics["accuracy"],
        "precision": metrics["precision"],
        "recall":    metrics["recall"],
    }


# Resume / Resume.

def load_resume_state(
    resume_path: str,
    model:       nn.Module,
    optimizer:   torch.optim.Optimizer,
    device:      str,
) -> dict:
    print(f"[EFFB0] Resuming from {resume_path}")
    ckpt = torch.load(resume_path, map_location=device, weights_only=False)

    model.load_state_dict(ckpt["model"])
    optimizer.load_state_dict(ckpt["optimizer"])

    start_epoch = int(ckpt["epoch"]) + 1
    print(f"[EFFB0] Resumed at epoch {start_epoch} "
          f"(checkpoint saved at epoch {ckpt['epoch']}).")
    return {
        "start_epoch":        start_epoch,
        "checkpoint_classes": ckpt.get("class_names", []),
    }


def load_best_state(best_path: Path) -> dict:
    if not best_path.exists():
        return {"best_acc": -1.0, "best_epoch": -1}
    bckpt = torch.load(best_path, map_location="cpu", weights_only=False)
    return {
        "best_acc":   float(bckpt.get("metric", -1.0)),
        "best_epoch": int(bckpt.get("epoch", -1)),
    }


# Main / Main.

def train(data_dir: str, cfg: Optional[dict] = None) -> None:
    c = {**DEFAULT_CFG, **(cfg or {})}

    set_seed(c["SEED"])

    if c["DEVICE"].startswith("cuda") and not torch.cuda.is_available():
        print("[EFFB0] CUDA not available, falling back to CPU.")
        c["DEVICE"] = "cpu"
        c["USE_AMP"] = False

    is_resume = c.get("RESUME_FROM") is not None

    print(f"[EFFB0] Model:    EfficientNet-B0 classifier")
    print(f"[EFFB0] Device:   {c['DEVICE']}")
    print(f"[EFFB0] Batch:    {c['BATCH_SIZE']}, Epochs: {c['EPOCHS']}")
    print(f"[EFFB0] LR0:      {c['LR0']}, LRF: {c['LRF']}")
    print(f"[EFFB0] AMP:      {c['USE_AMP']}")
    print(f"[EFFB0] Patience: {c['PATIENCE']}")
    if is_resume:
        print(f"[EFFB0] Resume:   {c['RESUME_FROM']}")

    train_loader, val_loader, class_names = build_dataloaders(data_dir, c)
    num_classes = len(class_names)
    print(f"[EFFB0] Train: {len(train_loader.dataset)} crops, {len(train_loader)} batches")
    print(f"[EFFB0] Val:   {len(val_loader.dataset)} crops, {len(val_loader)} batches")
    print(f"[EFFB0] Classes ({num_classes}): {class_names}")

    model = create_efficientnet_b0(
        num_classes=num_classes,
        pretrained=c["PRETRAINED"],
        pretrained_path=c["PRETRAINED_PATH"],
    ).to(c["DEVICE"])

    params = count_parameters(model)
    print(f"[EFFB0] Total params: {params['total_millions']:.2f}M")

    optimizer = build_optimizer(model, c)
    criterion = nn.CrossEntropyLoss()

    use_amp = c["USE_AMP"] and c["DEVICE"].startswith("cuda")
    scaler  = GradScaler("cuda") if use_amp else None
    if use_amp:
        torch.backends.cudnn.benchmark = True

    start_epoch      = 0
    best_acc         = -1.0
    best_epoch       = -1
    patience_counter = 0

    if is_resume:
        save_dir   = Path(c["RESUME_FROM"]).resolve().parent
        excel_path = save_dir / c["OUTPUT_XLSX"]
        if not excel_path.exists():
            raise FileNotFoundError(
                f"Excel not found in resume dir: {excel_path}. "
                f"Cannot resume without existing metrics file."
            )
        print(f"[EFFB0] Resume dir: {save_dir}")

        resume_state = load_resume_state(c["RESUME_FROM"], model, optimizer, c["DEVICE"])
        start_epoch  = resume_state["start_epoch"]

        best_state       = load_best_state(save_dir / "best.pt")
        best_acc         = best_state["best_acc"]
        best_epoch       = best_state["best_epoch"]
        patience_counter = max(0, (start_epoch - 1) - best_epoch)
        print(f"[EFFB0] Best so far: val_acc={best_acc:.4f} @ epoch {best_epoch}")
        print(f"[EFFB0] Patience counter: {patience_counter}/{c['PATIENCE']}")
    else:
        save_dir = resolve_run_dir(c["PROJECT"], c["NAME"])
        save_dir.mkdir(parents=True, exist_ok=True)
        excel_path = save_dir / c["OUTPUT_XLSX"]
        print(f"[EFFB0] Output dir: {save_dir}")
        build_excel(excel_path)

    print(f"\n[EFFB0] Starting training from epoch {start_epoch}...\n")

    for epoch in range(start_epoch, c["EPOCHS"]):
        epoch_start = time.time()

        base_lr = compute_lr(
            epoch=epoch,
            total_epochs=c["EPOCHS"],
            warmup_epochs=c["WARMUP_EPOCHS"],
            lr0=c["LR0"],
            lrf=c["LRF"],
        )
        set_param_group_lrs(optimizer, base_lr, c)

        print(f"=== Epoch {epoch}/{c['EPOCHS'] - 1}  base_lr={base_lr:.6f} ===")

        train_metrics = train_one_epoch(
            model=model,
            optimizer=optimizer,
            criterion=criterion,
            loader=train_loader,
            device=c["DEVICE"],
            epoch=epoch,
            scaler=scaler,
            use_amp=use_amp,
            log_interval=c["LOG_INTERVAL"],
            num_classes=num_classes,
        )
        print(f"  Train: loss={train_metrics['loss']:.4f}  "
              f"acc={train_metrics['accuracy']:.4f}  "
              f"P={train_metrics['precision']:.4f}  "
              f"R={train_metrics['recall']:.4f}")

        val_metrics = validate(
            model=model,
            criterion=criterion,
            loader=val_loader,
            device=c["DEVICE"],
            num_classes=num_classes,
        )
        print(f"  Val:   loss={val_metrics['loss']:.4f}  "
              f"acc={val_metrics['accuracy']:.4f}  "
              f"P={val_metrics['precision']:.4f}  "
              f"R={val_metrics['recall']:.4f}")

        append_epoch_row(
            excel_path,
            [
                epoch,
                round(base_lr, 8),
                _fmt(train_metrics["loss"]),
                _fmt(train_metrics["accuracy"]),
                _fmt(train_metrics["precision"]),
                _fmt(train_metrics["recall"]),
                _fmt(val_metrics["loss"]),
                _fmt(val_metrics["accuracy"]),
                _fmt(val_metrics["precision"]),
                _fmt(val_metrics["recall"]),
            ],
        )

        torch.save(
            {
                "epoch":       epoch,
                "model":       model.state_dict(),
                "optimizer":   optimizer.state_dict(),
                "cfg":         c,
                "class_names": class_names,
            },
            save_dir / "last.pt",
        )

        if val_metrics["accuracy"] > best_acc:
            best_acc         = val_metrics["accuracy"]
            best_epoch       = epoch
            patience_counter = 0
            torch.save(
                {
                    "epoch":       epoch,
                    "model":       model.state_dict(),
                    "optimizer":   optimizer.state_dict(),
                    "cfg":         c,
                    "class_names": class_names,
                    "metric":      val_metrics["accuracy"],
                },
                save_dir / "best.pt",
            )
            print(f"  >> NEW BEST val_acc={best_acc:.4f} at epoch {best_epoch}")
        else:
            patience_counter += 1
            print(f"  No improvement ({patience_counter}/{c['PATIENCE']})  "
                  f"best val_acc={best_acc:.4f} @ epoch {best_epoch}")

        epoch_time = time.time() - epoch_start
        print(f"  Epoch time: {epoch_time:.1f}s\n")

        if patience_counter >= c["PATIENCE"]:
            print(f"[EFFB0] Early stopping at epoch {epoch}. "
                  f"Best: epoch {best_epoch}, val_acc={best_acc:.4f}")
            break

    print(f"\n[EFFB0] Training finished.")
    print(f"  Best epoch:    {best_epoch}")
    print(f"  Best val_acc:  {best_acc:.4f}")
    print(f"  Output dir:    {save_dir}")
