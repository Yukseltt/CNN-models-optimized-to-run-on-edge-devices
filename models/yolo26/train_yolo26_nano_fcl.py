# train_yolo26_nano_fcl.py
# YOLO26-nano training with Focal Loss and Excel metric logging.

import sys
from pathlib import Path
import argparse
from typing import Optional

# 1. Proje ana dizinini bul ve sisteme ekle (Hata onleme)
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import yaml
import torch
import torch.nn as nn
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ultralytics import YOLO
from ultralytics.utils import LOGGER

# 2. Artik guvenle import edebilir
from src.focal_loss_func import binary_focal_loss_with_logits

RUNS_DIR = PROJECT_ROOT / "runs"

NANO_CFG = {
    "IMG_SIZE":       640,
    "DEVICE":         0,
    "EPOCHS":         200, # Kullanici istegine gore 100 olarak guncellendi
    "BATCH_SIZE":     32,
    "PATIENCE":       30,
    "WORKERS":        4,
    "CACHE":          False,
    "SEED":           0,
    "OPTIMIZER":      "SGD",
    "LR0":            0.01,
    "LRF":            0.1,
    "MOMENTUM":       0.937,
    "WEIGHT_DECAY":   0.0005,
    "WARMUP_EPOCHS":  3,
    "AMP":            True,
    "HSV_H":          0.015,
    "HSV_S":          0.7,
    "HSV_V":          0.4,
    "MOSAIC":         0.0,
    "MIXUP":          0.0,
    "FLIPUD":         0.0,
    "FLIPLR":         0.0,
    "DEGREES":        0.0,
    "TRANSLATE":      0.0,
    "SCALE":          0.0,
    "PERSPECTIVE":    0.0,
    "ERASING":        0.0,
    "FCL_ALPHA":      0.25,
    "FCL_GAMMA":      2.0,
    "PROJECT":        str(RUNS_DIR),
    "NAME":           "yolo26n_fcl_default",
    "OUTPUT_XLSX":    "training_metrics.xlsx",
}

class FocalBCELoss(nn.Module):
    def __init__(self, alpha: float = 0.25, gamma: float = 2.0):
        super().__init__()
        self.alpha = alpha
        self.gamma = gamma

    def forward(self, pred: torch.Tensor, target: torch.Tensor) -> torch.Tensor:
        return binary_focal_loss_with_logits(
            logits=pred, targets=target, alpha=self.alpha, gamma=self.gamma, reduction="none",
        )

def read_class_names(data_yaml: str) -> list[str]:
    with open(data_yaml, "r") as f:
        data = yaml.safe_load(f)
    names = data.get("names")
    if names is None: raise ValueError(f"'names' field not found in {data_yaml}")
    if isinstance(names, dict): return [names[k] for k in sorted(names.keys())]
    elif isinstance(names, list): return list(names)
    else: raise ValueError(f"'names' must be list or dict, got: {type(names)}")

def resolve_excel_path(output: str) -> Path:
    path = Path(output)
    return path if path.is_absolute() else RUNS_DIR / path

STATIC_HEADERS = [
    "Epoch", "Learning Rate", "Box Loss (Train)", "Box Loss (Val)",
    "Class Loss (Train)", "Class Loss (Val)", "DFL Loss (Train)", "DFL Loss (Val)",
    "Precision (Val)", "Recall (Val)", "F1 (Val)", "mAP@0.5 (overall)", "mAP@0.5:0.95 (overall)",
]

_THIN        = Side(border_style="thin", color="BBBBBB")
BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)

def _build_headers(class_names: list[str]) -> list[str]:
    headers = list(STATIC_HEADERS)
    for name in class_names:
        headers.extend([f"mAP@0.5 ({name})", f"mAP@0.5:0.95 ({name})", f"Precision ({name})", f"Recall ({name})"])
    return headers

def _apply_header(ws, class_names: list[str]) -> None:
    headers = _build_headers(class_names)
    for ci, val in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10 if i <= len(STATIC_HEADERS) else 18
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

def build_excel(output_path: Path, class_names: list[str]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Training Metrics"
    _apply_header(ws, class_names)
    cm = wb.create_sheet("Confusion Matrix")
    cm["A1"] = "Confusion Matrix - Egitim tamamlandiktan sonra doldurulur."
    cm["A1"].font = Font(name="Arial", bold=True, size=11)
    ch = wb.create_sheet("Charts")
    for r, txt in enumerate(["Grafikler bu sayfaya eklenecektir.", "Insert -> Chart ile Training Metrics verilerini kullanabilirsiniz."], start=1):
        c = ch.cell(row=r, column=1, value=txt)
        c.font = Font(name="Arial", size=10, italic=True, color="555555")
    wb.save(output_path)
    LOGGER.info(f"[FCL-Train] Excel created: {output_path}")

def _fmt(v) -> Optional[float]: return None if v is None else round(float(v), 4)

def append_epoch_row(output_path: Path, row_values: list) -> None:
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Training Metrics"]
    data_row = ws.max_row + 1
    alt_fill = (PatternFill("solid", start_color="EBF3FB") if data_row % 2 == 0 else None)
    for ci, val in enumerate(row_values, 1):
        c = ws.cell(row=data_row, column=ci, value=val)
        c.font, c.border, c.alignment = DATA_FONT, BORDER, Alignment(horizontal="center")
        if alt_fill: c.fill = alt_fill
    wb.save(output_path)

class FCLTrainingCallbacks:
    def __init__(self, output_xlsx_name: str, class_names: list[str], alpha: float, gamma: float):
        self.output_xlsx_name = output_xlsx_name
        self.class_names, self.alpha, self.gamma = class_names, alpha, gamma
        self.output_path, self.injected = None, False

    def on_pretrain_routine_end(self, trainer) -> None:
        self.output_path = Path(trainer.save_dir) / self.output_xlsx_name
        if not self.output_path.exists(): build_excel(self.output_path, self.class_names)

    def on_train_start(self, trainer) -> None:
        criterion = getattr(trainer, "criterion", None)
        if criterion is None or not hasattr(criterion, "bce"):
            LOGGER.warning("[FCL-Train] Cannot inject focal loss.")
            return
        try: device = next(trainer.model.parameters()).device
        except Exception: device = torch.device("cpu")
        criterion.bce = FocalBCELoss(alpha=self.alpha, gamma=self.gamma).to(device)
        self.injected = True
        LOGGER.info(f"[FCL-Train] Focal Loss injected (alpha={self.alpha}, gamma={self.gamma})")

    def on_fit_epoch_end(self, trainer) -> None:
        epoch, metrics, loss_items = trainer.epoch, trainer.metrics, trainer.loss_items
        lr_list = trainer.scheduler.get_last_lr()
        lr_val = lr_list[0] if lr_list else trainer.args.lr0

        def _li(i): return float(loss_items[i]) if loss_items is not None else None
        box_t, cls_t, dfl_t = _li(0), _li(1), _li(2)

        def _m(key): return float(metrics[key]) if key in metrics else None
        box_v, cls_v, dfl_v = _m("val/box_loss"), _m("val/cls_loss"), _m("val/dfl_loss")
        prec, rec = _m("metrics/precision(B)"), _m("metrics/recall(B)")
        map50, map50_95 = _m("metrics/mAP50(B)"), _m("metrics/mAP50-95(B)")

        f1 = (2 * prec * rec / (prec + rec)) if prec and rec and (prec + rec) > 0 else 0.0

        per_class = self._collect_per_class(trainer)
        row = [(epoch + 1), round(lr_val, 8), _fmt(box_t), _fmt(box_v), _fmt(cls_t), _fmt(cls_v), _fmt(dfl_t), _fmt(dfl_v), _fmt(prec), _fmt(rec), _fmt(f1), _fmt(map50), _fmt(map50_95)]
        for i in range(len(self.class_names)):
            row.extend([_fmt(per_class["map50"][i]), _fmt(per_class["map50_95"][i]), _fmt(per_class["precision"][i]), _fmt(per_class["recall"][i])])

        append_epoch_row(self.output_path, row)
        if all(v is not None for v in [box_t, cls_t, dfl_t, map50]):
            LOGGER.info(f"[FCL-Train] Epoch {epoch:3d} -> box={box_t:.4f} cls={cls_t:.4f} dfl={dfl_t:.4f} mAP@0.5={map50:.4f}")

    def _collect_per_class(self, trainer) -> dict:
        nc = len(self.class_names)
        res = {"map50": [None]*nc, "map50_95": [None]*nc, "precision": [None]*nc, "recall": [None]*nc}
        try:
            validator = getattr(trainer, "validator", None)
            if validator is None: return res
            v_metrics = getattr(validator, "metrics", None)
            if v_metrics is None: return res
            box = getattr(v_metrics, "box", None)
            if box is None: return res
            ap_class_index = getattr(box, "ap_class_index", None)
            if ap_class_index is None or len(ap_class_index) == 0: return res
            ap50_arr = getattr(box, "ap50", None)
            maps     = getattr(box, "maps", None)
            p_arr    = getattr(box, "p", None)
            # box.r per-class recall array; box.recall scalar olabilir, once r dene
            r_arr = getattr(box, "r", None)
            if r_arr is None: r_arr = getattr(box, "recall", None)
            if r_arr is not None and not hasattr(r_arr, "__len__"): r_arr = None
            for pos, cls_idx in enumerate(ap_class_index):
                cls_idx = int(cls_idx)
                if cls_idx >= nc: continue
                if ap50_arr is not None and pos < len(ap50_arr): res["map50"][cls_idx] = ap50_arr[pos]
                if maps is not None and cls_idx < len(maps): res["map50_95"][cls_idx] = maps[cls_idx]
                if p_arr is not None and pos < len(p_arr): res["precision"][cls_idx] = p_arr[pos]
                if r_arr is not None and pos < len(r_arr): res["recall"][cls_idx] = r_arr[pos]
        except Exception: pass
        return res

    def on_train_end(self, trainer) -> None:
        LOGGER.info(f"[FCL-Train] Training finished. Output: {self.output_path}")

def train(data: str, cfg: Optional[dict] = None, class_names: Optional[list[str]] = None) -> None:
    c = {**NANO_CFG, **(cfg or {})}
    if class_names is None: class_names = read_class_names(data)
    excel_path = resolve_excel_path(c["OUTPUT_XLSX"])

    LOGGER.info(f"[FCL-Train] Classes: {class_names}, alpha={c['FCL_ALPHA']}, gamma={c['FCL_GAMMA']}, seed={c['SEED']}")
    
    model = YOLO("yolo26n.pt")
    cbs = FCLTrainingCallbacks(output_xlsx_name=Path(c["OUTPUT_XLSX"]).name, class_names=class_names, alpha=c["FCL_ALPHA"], gamma=c["FCL_GAMMA"])
    
    model.add_callback("on_pretrain_routine_end", cbs.on_pretrain_routine_end)
    model.add_callback("on_train_start",          cbs.on_train_start)
    model.add_callback("on_fit_epoch_end",        cbs.on_fit_epoch_end)
    model.add_callback("on_train_end",            cbs.on_train_end)

    model.train(
        data=data, imgsz=c["IMG_SIZE"], device=c["DEVICE"], epochs=c["EPOCHS"],
        batch=c["BATCH_SIZE"], patience=c["PATIENCE"], workers=c["WORKERS"], cache=c["CACHE"],
        seed=c["SEED"], optimizer=c["OPTIMIZER"], lr0=c["LR0"], lrf=c["LRF"],
        momentum=c["MOMENTUM"], weight_decay=c["WEIGHT_DECAY"], warmup_epochs=c["WARMUP_EPOCHS"],
        amp=c["AMP"], hsv_h=c["HSV_H"], hsv_s=c["HSV_S"], hsv_v=c["HSV_V"], mosaic=c["MOSAIC"],
        mixup=c["MIXUP"], flipud=c["FLIPUD"], fliplr=c["FLIPLR"], degrees=c["DEGREES"],
        translate=c["TRANSLATE"], scale=c["SCALE"], perspective=c["PERSPECTIVE"],
        erasing=c["ERASING"], project=c["PROJECT"], name=c["NAME"], verbose=True,
    )