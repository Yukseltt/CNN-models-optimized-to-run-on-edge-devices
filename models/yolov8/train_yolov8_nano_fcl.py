# train_yolov8_nano_fcl.py
#
# YOLOv8-nano training with Focal Loss and Excel metric logging.
# YOLOv8-nano egitimi, Focal Loss ve Excel metrik kaydi ile.
#
# Install / Kurulum:
#     pip install ultralytics openpyxl pyyaml
#
# Usage / Kullanim:
#     python run_training.py
#
# Notes / Notlar:
#     focal_loss_func.py must be importable as src.focal_loss_func.
#     focal_loss_func.py src.focal_loss_func olarak import edilebilmelidir.
#
#     Focal Loss replaces the BCE in v8DetectionLoss.bce attribute.
#     Focal Loss, v8DetectionLoss.bce attribute'unu degistirir.
#
#     Per-class metrics are logged for each class defined in data.yaml.
#     data.yaml'da tanimli her sinif icin per-class metrikler loglanir.

import argparse
from pathlib import Path
from typing import Optional

import yaml
import torch
import torch.nn as nn
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ultralytics import YOLO
from ultralytics.utils import LOGGER

from src.focal_loss_func import binary_focal_loss_with_logits


# Project root directory (two levels up from this script).
# Proje kok dizini (bu scriptin iki ust klasoru).
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Default output directory for training artifacts and Excel logs.
# Egitim ciktilari ve Excel kayitlari icin varsayilan cikti dizini.
RUNS_DIR = PROJECT_ROOT / "runs"


# Training configuration with YOLOv8-nano defaults.
# YOLOv8-nano varsayilan degerleriyle egitim konfigurasyonu.

NANO_CFG = {
    "IMG_SIZE":       640,
    "DEVICE":         0,
    "EPOCHS":         300,
    "BATCH_SIZE":     32,
    "PATIENCE":       30,
    "WORKERS":        4,
    "CACHE":          False,
    "SEED":           0,

    "OPTIMIZER":      "SGD",
    "LR0":            0.01,
    "LRF":            0.1,
    "MOMENTUM":       0.937,
    "WEIGHT_DECAY":   0.0005,
    "WARMUP_EPOCHS":  3,
    "AMP":            True,

    "HSV_H":          0.015,
    "HSV_S":          0.7,
    "HSV_V":          0.4,
    "MOSAIC":         0.0,
    "MIXUP":          0.0,
    "FLIPUD":         0.0,
    "FLIPLR":         0.0,
    "DEGREES":        0.0,
    "TRANSLATE":      0.0,
    "SCALE":          0.0,
    "PERSPECTIVE":    0.0,
    "ERASING":        0.0,

    # Focal Loss hyperparameters.
    # Focal Loss hyperparametreleri.
    "FCL_ALPHA":      0.25,
    "FCL_GAMMA":      2.0,

    "PROJECT":        str(RUNS_DIR),
    "NAME":           "yolov8n_fcl_default",
    "OUTPUT_XLSX":    "training_metrics.xlsx",
}


# Focal BCE wrapper class.
# Focal BCE sarmalayici sinif.

class FocalBCELoss(nn.Module):
    # A drop-in replacement for nn.BCEWithLogitsLoss(reduction="none").
    # nn.BCEWithLogitsLoss(reduction="none") yerine direkt kullanilan sarmalayici.
    #
    # v8DetectionLoss expects self.bce(pred, target) to return an element-wise
    # loss tensor with the same shape as pred. We honor that contract.
    # v8DetectionLoss self.bce(pred, target) cagrisinin pred ile ayni shape'de
    # element-wise loss tensor dondurmesini bekler. Biz bu sozlesmeyi koruyoruz.

    def __init__(self, alpha: float = 0.25, gamma: float = 2.0):
        super().__init__()
        self.alpha = alpha
        self.gamma = gamma

    def forward(self, pred: torch.Tensor, target: torch.Tensor) -> torch.Tensor:
        # Same signature as nn.BCEWithLogitsLoss(reduction="none").
        # nn.BCEWithLogitsLoss(reduction="none") ile ayni imza.
        return binary_focal_loss_with_logits(
            logits=pred,
            targets=target,
            alpha=self.alpha,
            gamma=self.gamma,
            reduction="none",
        )


# Read class names from data.yaml.
# data.yaml'dan sinif isimlerini okur.

def read_class_names(data_yaml: str) -> list[str]:
    # Load class names from the data.yaml file.
    # data.yaml dosyasindan sinif isimlerini yukler.
    with open(data_yaml, "r") as f:
        data = yaml.safe_load(f)
    names = data.get("names")
    if names is None:
        raise ValueError(f"'names' field not found in {data_yaml}")
    if isinstance(names, dict):
        return [names[k] for k in sorted(names.keys())]
    elif isinstance(names, list):
        return list(names)
    else:
        raise ValueError(f"'names' must be list or dict, got: {type(names)}")


# Resolve Excel output path.
# Excel cikti yolunu cozer.

def resolve_excel_path(output: str) -> Path:
    path = Path(output)
    if path.is_absolute():
        return path
    return RUNS_DIR / path


# Excel template creation and epoch row appending.
# Excel sablonu olusturma ve epoch satiri ekleme.

STATIC_HEADERS = [
    "Epoch", "Learning Rate",
    "Box Loss (Train)", "Box Loss (Val)",
    "Class Loss (Train)", "Class Loss (Val)",
    "DFL Loss (Train)", "DFL Loss (Val)",
    "Precision (Val)", "Recall (Val)", "F1 (Val)",
    "mAP@0.5 (overall)", "mAP@0.5:0.95 (overall)",
]

_THIN        = Side(border_style="thin", color="BBBBBB")
BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)


def _build_headers(class_names: list[str]) -> list[str]:
    # Build the full header row based on class names.
    # Sinif isimlerine gore tam header satirini olusturur.
    headers = list(STATIC_HEADERS)
    for name in class_names:
        headers.append(f"mAP@0.5 ({name})")
        headers.append(f"mAP@0.5:0.95 ({name})")
        headers.append(f"Precision ({name})")
        headers.append(f"Recall ({name})")
    return headers


def _apply_header(ws, class_names: list[str]) -> None:
    # Apply single-row header to the Training Metrics sheet.
    # Training Metrics sayfasina tek satirli header uygular.
    headers = _build_headers(class_names)
    for ci, val in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for i in range(1, len(headers) + 1):
        width = 10 if i <= len(STATIC_HEADERS) else 18
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"


def build_excel(output_path: Path, class_names: list[str]) -> None:
    # Create a new Excel file with three sheets.
    # Uc sayfali yeni bir Excel dosyasi olusturur.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Training Metrics"
    _apply_header(ws, class_names)

    cm = wb.create_sheet("Confusion Matrix")
    cm["A1"] = "Confusion Matrix - Egitim tamamlandiktan sonra doldurulur."
    cm["A1"].font = Font(name="Arial", bold=True, size=11)

    ch = wb.create_sheet("Charts")
    for r, txt in enumerate([
        "Grafikler bu sayfaya eklenecektir.",
        "Insert -> Chart ile Training Metrics verilerini kullanabilirsiniz.",
    ], start=1):
        c = ch.cell(row=r, column=1, value=txt)
        c.font = Font(name="Arial", size=10, italic=True, color="555555")

    wb.save(output_path)
    LOGGER.info(f"[FCL-Train] Excel created: {output_path}")


def _fmt(v) -> Optional[float]:
    # Preserve None, round float to 4 decimals.
    # None'i korur, float degerini 4 ondalikla yuvarlar.
    return None if v is None else round(float(v), 4)


def append_epoch_row(output_path: Path, row_values: list) -> None:
    # Append a single epoch row to the existing Excel file.
    # Mevcut Excel dosyasina tek bir epoch satiri ekler.
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Training Metrics"]
    data_row = ws.max_row + 1
    alt_fill = (PatternFill("solid", start_color="EBF3FB")
                if data_row % 2 == 0 else None)
    for ci, val in enumerate(row_values, 1):
        c = ws.cell(row=data_row, column=ci, value=val)
        c.font = DATA_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill
    wb.save(output_path)


# Trainer callbacks for Focal Loss injection and metric logging.
# Focal Loss inject ve metrik kaydi icin trainer callback'leri.

class FCLTrainingCallbacks:
    # Callback class attached to the YOLO trainer.
    # YOLO trainer'ina eklenen callback sinifi.
    #
    # Hooks:
    #     on_train_start: replace v8DetectionLoss.bce with FocalBCELoss.
    #     on_train_start: v8DetectionLoss.bce'yi FocalBCELoss ile degistirir.
    #
    #     on_pretrain_routine_end: build Excel inside the actual run dir.
    #     on_pretrain_routine_end: Excel'i gercek run dizini icinde olusturur.
    #
    #     on_fit_epoch_end: collect metrics and write to Excel.
    #     on_fit_epoch_end: metrikleri toplar ve Excel'e yazar.
    #
    #     on_train_end: log training completion.
    #     on_train_end: egitimin tamamlandigini loglar.

    def __init__(
        self,
        output_xlsx_name: str,
        class_names:      list[str],
        alpha:            float,
        gamma:            float,
    ):
        self.output_xlsx_name = output_xlsx_name
        self.class_names      = class_names
        self.alpha            = alpha
        self.gamma            = gamma
        self.output_path      = None
        self.injected         = False

    def on_pretrain_routine_end(self, trainer) -> None:
        # Build Excel in the actual YOLO save_dir.
        # Excel'i gercek YOLO save_dir'i icinde olustur.
        save_dir = Path(trainer.save_dir)
        self.output_path = save_dir / self.output_xlsx_name
        if not self.output_path.exists():
            build_excel(self.output_path, self.class_names)
        LOGGER.info(f"[FCL-Train] Excel will be written to: {self.output_path}")

    def on_train_start(self, trainer) -> None:
        # Replace v8DetectionLoss.bce with FocalBCELoss.
        # v8DetectionLoss.bce'yi FocalBCELoss ile degistir.
        #
        # By this hook, criterion is already initialized.
        # Bu hook noktasinda criterion zaten init olmus durumda.
        # We mutate the existing instance instead of replacing the class.
        # Sinifi degistirmek yerine mevcut instance'i mutate ediyoruz.
        criterion = getattr(trainer, "criterion", None)
        if criterion is None:
            LOGGER.warning("[FCL-Train] trainer.criterion is None, cannot inject focal loss.")
            return

        if not hasattr(criterion, "bce"):
            LOGGER.warning("[FCL-Train] criterion has no .bce attribute, cannot inject focal loss.")
            return

        # Determine device from the existing bce attribute (or model).
        # Mevcut bce attribute'undan (veya modelden) device'i belirle.
        try:
            device = next(trainer.model.parameters()).device
        except Exception:
            device = torch.device("cpu")

        focal = FocalBCELoss(alpha=self.alpha, gamma=self.gamma).to(device)
        criterion.bce = focal
        self.injected = True

        LOGGER.info(
            f"[FCL-Train] Focal Loss injected into criterion.bce "
            f"(alpha={self.alpha}, gamma={self.gamma})"
        )

    def on_fit_epoch_end(self, trainer) -> None:
        # Collect metrics at the end of an epoch and write them to Excel.
        # Epoch sonunda metrikleri toplar ve Excel'e yazar.
        epoch      = trainer.epoch
        metrics    = trainer.metrics
        loss_items = trainer.loss_items
        lr_list    = trainer.scheduler.get_last_lr()
        lr_val     = lr_list[0] if lr_list else trainer.args.lr0

        def _li(i):
            return float(loss_items[i]) if loss_items is not None else None

        box_t, cls_t, dfl_t = _li(0), _li(1), _li(2)

        def _m(key):
            return float(metrics[key]) if key in metrics else None

        box_v    = _m("val/box_loss")
        cls_v    = _m("val/cls_loss")
        dfl_v    = _m("val/dfl_loss")
        prec     = _m("metrics/precision(B)")
        rec      = _m("metrics/recall(B)")
        map50    = _m("metrics/mAP50(B)")
        map50_95 = _m("metrics/mAP50-95(B)")

        f1 = None
        if prec is not None and rec is not None:
            denom = prec + rec
            f1 = (2 * prec * rec / denom) if denom > 0 else 0.0

        per_class = self._collect_per_class(trainer)

        row = [
            epoch, round(lr_val, 8),
            _fmt(box_t), _fmt(box_v),
            _fmt(cls_t), _fmt(cls_v),
            _fmt(dfl_t), _fmt(dfl_v),
            _fmt(prec), _fmt(rec), _fmt(f1),
            _fmt(map50), _fmt(map50_95),
        ]
        for i in range(len(self.class_names)):
            row.extend([
                _fmt(per_class["map50"][i]),
                _fmt(per_class["map50_95"][i]),
                _fmt(per_class["precision"][i]),
                _fmt(per_class["recall"][i]),
            ])

        append_epoch_row(self.output_path, row)

        if all(v is not None for v in [box_t, cls_t, dfl_t, map50]):
            LOGGER.info(
                f"[FCL-Train] Epoch {epoch:3d} -> "
                f"box={box_t:.4f}  cls={cls_t:.4f}  dfl={dfl_t:.4f}  "
                f"mAP@0.5={map50:.4f}"
            )
        else:
            LOGGER.info(f"[FCL-Train] Epoch {epoch} saved.")

    def _collect_per_class(self, trainer) -> dict:
        # Collect per-class metrics from trainer.validator.metrics.box.
        # trainer.validator.metrics.box'tan per-class metrikleri toplar.
        nc = len(self.class_names)
        result = {
            "map50":     [None] * nc,
            "map50_95":  [None] * nc,
            "precision": [None] * nc,
            "recall":    [None] * nc,
        }
        validator = getattr(trainer, "validator", None)
        if validator is None:
            return result
        v_metrics = getattr(validator, "metrics", None)
        if v_metrics is None:
            return result
        box = getattr(v_metrics, "box", None)
        if box is None:
            return result
        ap_class_index = getattr(box, "ap_class_index", None)
        if ap_class_index is None or len(ap_class_index) == 0:
            return result
        maps     = getattr(box, "maps", None)
        p_arr    = getattr(box, "p", None)
        r_arr    = getattr(box, "recall", None)
        if r_arr is None:
            r_arr = getattr(box, "r", None)
        ap50_arr = getattr(box, "ap50", None)
        for pos, cls_idx in enumerate(ap_class_index):
            cls_idx = int(cls_idx)
            if cls_idx >= nc:
                continue
            if ap50_arr is not None and pos < len(ap50_arr):
                result["map50"][cls_idx] = ap50_arr[pos]
            if maps is not None and cls_idx < len(maps):
                result["map50_95"][cls_idx] = maps[cls_idx]
            if p_arr is not None and pos < len(p_arr):
                result["precision"][cls_idx] = p_arr[pos]
            if r_arr is not None and pos < len(r_arr):
                result["recall"][cls_idx] = r_arr[pos]
        return result

    def on_train_end(self, trainer) -> None:
        LOGGER.info(
            f"[FCL-Train] Training finished. Output: {self.output_path}"
        )


# Main training function.
# Ana egitim fonksiyonu.

def train(
    data:         str,
    cfg:          Optional[dict] = None,
    class_names:  Optional[list[str]] = None,
) -> None:
    # Start YOLOv8-nano training with Focal Loss.
    # YOLOv8-nano egitimini Focal Loss ile baslatir.
    c = {**NANO_CFG, **(cfg or {})}

    if class_names is None:
        class_names = read_class_names(data)

    excel_path = resolve_excel_path(c["OUTPUT_XLSX"])

    LOGGER.info(
        f"[FCL-Train] Classes: {class_names}, "
        f"alpha={c['FCL_ALPHA']}, gamma={c['FCL_GAMMA']}, seed={c['SEED']}"
    )
    LOGGER.info(f"[FCL-Train] Runs directory: {c['PROJECT']}")

    model = YOLO("yolov8n.pt")

    cbs = FCLTrainingCallbacks(
        output_xlsx_name=Path(c["OUTPUT_XLSX"]).name,
        class_names=class_names,
        alpha=c["FCL_ALPHA"],
        gamma=c["FCL_GAMMA"],
    )

    model.add_callback("on_pretrain_routine_end", cbs.on_pretrain_routine_end)
    model.add_callback("on_train_start",          cbs.on_train_start)
    model.add_callback("on_fit_epoch_end",        cbs.on_fit_epoch_end)
    model.add_callback("on_train_end",            cbs.on_train_end)

    model.train(
        data=data,
        imgsz=c["IMG_SIZE"],
        device=c["DEVICE"],
        epochs=c["EPOCHS"],
        batch=c["BATCH_SIZE"],
        patience=c["PATIENCE"],
        workers=c["WORKERS"],
        cache=c["CACHE"],
        seed=c["SEED"],
        optimizer=c["OPTIMIZER"],
        lr0=c["LR0"],
        lrf=c["LRF"],
        momentum=c["MOMENTUM"],
        weight_decay=c["WEIGHT_DECAY"],
        warmup_epochs=c["WARMUP_EPOCHS"],
        amp=c["AMP"],
        hsv_h=c["HSV_H"],
        hsv_s=c["HSV_S"],
        hsv_v=c["HSV_V"],
        mosaic=c["MOSAIC"],
        mixup=c["MIXUP"],
        flipud=c["FLIPUD"],
        fliplr=c["FLIPLR"],
        degrees=c["DEGREES"],
        translate=c["TRANSLATE"],
        scale=c["SCALE"],
        perspective=c["PERSPECTIVE"],
        erasing=c["ERASING"],
        project=c["PROJECT"],
        name=c["NAME"],
        verbose=True,
    )