# train_yolov8_nano_cb.py
#
# YOLOv8-nano training with Class-Balanced Loss and Excel metric logging.
# YOLOv8-nano egitimi, Class-Balanced Loss ve Excel metrik kaydi ile.
#
# Install / Kurulum:
#     pip install ultralytics openpyxl pyyaml
#
# Usage / Kullanim:
#     python run_training.py
#
# Notes / Notlar:
#     cb_loss.py must be importable as src.cb_loss from project root.
#     cb_loss.py proje kokunden src.cb_loss olarak import edilebilmelidir.
#     run_training.py adds the project root to sys.path.
#     run_training.py proje kokunu sys.path'e ekler.
#
#     CB weights are injected via model.class_weights into v8DetectionLoss.
#     CB agirliklari model.class_weights uzerinden v8DetectionLoss'a inject edilir.
#
#     Per-class metrics are logged for each class defined in data.yaml.
#     data.yaml'da tanimli her sinif icin per-class metrikler loglanir.

import argparse
from pathlib import Path
from typing import Optional

import yaml
import torch
import torch.nn as nn
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ultralytics import YOLO
from ultralytics.utils import LOGGER

from src.cb_loss import compute_cb_weights


# Project root directory (two levels up from this script).
# Proje kok dizini (bu scriptin iki ust klasoru).
# models/yolov8/train_yolov8_nano_cb.py -> uc_cihazlarda_terhmal_object_detection/
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Default output directory for training artifacts and Excel logs.
# Egitim ciktilari ve Excel kayitlari icin varsayilan cikti dizini.
RUNS_DIR = PROJECT_ROOT / "runs"


# Training configuration with YOLOv8-nano defaults.
# YOLOv8-nano varsayilan degerleriyle egitim konfigurasyonu.

NANO_CFG = {
    "IMG_SIZE":       640,
    "DEVICE":         0,
    "EPOCHS":         1000,
    "BATCH_SIZE":     32,
    "PATIENCE":       30,
    "WORKERS":        4,
    "CACHE":          False,
    "SEED":           0,

    "OPTIMIZER":      "SGD",
    "LR0":            0.01,
    "LRF":            0.1,
    "MOMENTUM":       0.937,
    "WEIGHT_DECAY":   0.0005,
    "WARMUP_EPOCHS":  10,
    "AMP":            True,

    "HSV_H":          0.015,
    "HSV_S":          0.7,
    "HSV_V":          0.4,
    "MOSAIC":         1.0,
    "MIXUP":          0.0,
    "FLIPUD":         0.0,
    "FLIPLR":         0.0,
    "DEGREES":        0.0,
    "TRANSLATE":      0.0,
    "SCALE":          0.0,
    "PERSPECTIVE":    0.0,
    "ERASING":        0.0,

    "BETA":           0.9999999,

    # Project is the parent directory for YOLO run folders.
    # Project YOLO run klasorlerinin ust dizinidir.
    "PROJECT":        str(RUNS_DIR),
    "NAME":           "yolov8n_cb_2",
    "OUTPUT_XLSX":    "training_results_2.xlsx",
}


# Read class names from data.yaml.
# data.yaml'dan sinif isimlerini okur.

def read_class_names(data_yaml: str) -> list[str]:
    # Load class names from the data.yaml file.
    # data.yaml dosyasindan sinif isimlerini yukler.
    #
    # Supports both list and dict formats for the 'names' field.
    # 'names' alani icin hem liste hem dict formatini destekler.
    with open(data_yaml, "r") as f:
        data = yaml.safe_load(f)

    names = data.get("names")
    if names is None:
        raise ValueError(f"'names' field not found in {data_yaml}")

    if isinstance(names, dict):
        # Sort by integer key to preserve class order.
        # Sinif sirasini korumak icin integer key'e gore sirala.
        return [names[k] for k in sorted(names.keys())]
    elif isinstance(names, list):
        return list(names)
    else:
        raise ValueError(f"'names' must be list or dict, got: {type(names)}")


# Resolve Excel output path.
# Excel cikti yolunu cozer.

def resolve_excel_path(output: str) -> Path:
    # If output is an absolute path, use it as is.
    # output mutlak yolsa oldugu gibi kullan.
    # Otherwise place it under the runs/ directory.
    # Aksi halde runs/ dizini altina yerlestir.
    path = Path(output)
    if path.is_absolute():
        return path
    return RUNS_DIR / path


# CB Loss integration via model.class_weights injection.
# model.class_weights uzerinden CB Loss entegrasyonu.

def inject_cb_weights(
    model,
    class_counts: list[int],
    beta: float,
    device: torch.device,
) -> None:
    # Set CB weights as model.class_weights attribute.
    # CB agirliklarini model.class_weights olarak set eder.
    #
    # v8DetectionLoss reads this attribute during init and multiplies BCE loss by it.
    # v8DetectionLoss init sirasinda bu attribute'u okur ve BCE loss ile carpar.
    weights = compute_cb_weights(
        class_counts=class_counts,
        beta=beta,
        normalize=True,
        device=device,
    )
    # Shape (1, 1, nc) for broadcasting over (bs, num_anchors, nc).
    # (bs, num_anchors, nc) uzerinde broadcast icin (1, 1, nc) shape'i.
    model.class_weights = weights.view(1, 1, -1)

    LOGGER.info(
        f"[CB-Train] class_weights injected - "
        f"beta={beta}, nc={len(class_counts)}, "
        f"min={weights.min():.4f}, max={weights.max():.4f}"
    )


# Excel template creation and epoch row appending.
# Excel sablonu olusturma ve epoch satiri ekleme.

STATIC_HEADERS = [
    "Epoch", "Learning Rate",
    "Box Loss (Train)", "Box Loss (Val)",
    "Class Loss (Train)", "Class Loss (Val)",
    "DFL Loss (Train)", "DFL Loss (Val)",
    "Precision (Val)", "Recall (Val)", "F1 (Val)",
    "mAP@0.5 (overall)", "mAP@0.5:0.95 (overall)",
]

_THIN        = Side(border_style="thin", color="BBBBBB")
BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)


def _build_headers(class_names: list[str]) -> list[str]:
    # Build the full header row based on class names.
    # Sinif isimlerine gore tam header satirini olusturur.
    #
    # For each class, 4 columns are added: mAP@0.5, mAP@0.5:0.95, Precision, Recall.
    # Her sinif icin 4 sutun eklenir: mAP@0.5, mAP@0.5:0.95, Precision, Recall.
    headers = list(STATIC_HEADERS)
    for name in class_names:
        headers.append(f"mAP@0.5 ({name})")
        headers.append(f"mAP@0.5:0.95 ({name})")
        headers.append(f"Precision ({name})")
        headers.append(f"Recall ({name})")
    return headers


def _apply_header(ws, class_names: list[str]) -> None:
    # Apply single-row header to the Training Metrics sheet.
    # Training Metrics sayfasina tek satirli header uygular.
    headers = _build_headers(class_names)

    for ci, val in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    # Static columns narrower, per-class columns wider for readability.
    # Static sutunlar dar, per-class sutunlar okunabilirlik icin genis.
    for i in range(1, len(headers) + 1):
        width = 10 if i <= len(STATIC_HEADERS) else 18
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"


def build_excel(output_path: Path, class_names: list[str]) -> None:
    # Create a new Excel file with three sheets.
    # Uc sayfali yeni bir Excel dosyasi olusturur.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Training Metrics"
    _apply_header(ws, class_names)

    cm = wb.create_sheet("Confusion Matrix")
    cm["A1"] = "Confusion Matrix - Egitim tamamlandiktan sonra doldurulur."
    cm["A1"].font = Font(name="Arial", bold=True, size=11)

    ch = wb.create_sheet("Charts")
    for r, txt in enumerate([
        "Grafikler bu sayfaya eklenecektir.",
        "Insert -> Chart ile Training Metrics verilerini kullanabilirsiniz.",
    ], start=1):
        c = ch.cell(row=r, column=1, value=txt)
        c.font = Font(name="Arial", size=10, italic=True, color="555555")

    wb.save(output_path)
    LOGGER.info(f"[CB-Train] Excel created: {output_path}")


def _fmt(v) -> Optional[float]:
    # Preserve None, round float to 4 decimals.
    # None'i korur, float degerini 4 ondalikla yuvarlar.
    return None if v is None else round(float(v), 4)


def append_epoch_row(
    output_path: Path,
    row_values: list,
) -> None:
    # Append a single epoch row to the existing Excel file.
    # Mevcut Excel dosyasina tek bir epoch satiri ekler.
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Training Metrics"]

    data_row = ws.max_row + 1

    alt_fill = (PatternFill("solid", start_color="EBF3FB")
                if data_row % 2 == 0 else None)

    for ci, val in enumerate(row_values, 1):
        c = ws.cell(row=data_row, column=ci, value=val)
        c.font = DATA_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill

    wb.save(output_path)


# Trainer callbacks for CB weight injection and metric logging.
# CB agirligi inject ve metrik kaydi icin trainer callback'leri.

class CBTrainingCallbacks:
    # Callback class attached to the YOLO trainer.
    # YOLO trainer'ina eklenen callback sinifi.
    #
    # Hooks:
    #     on_pretrain_routine_end: inject CB weights into the model.
    #     on_pretrain_routine_end: CB agirliklarini modele inject eder.
    #
    #     on_fit_epoch_end: collect metrics and write to Excel.
    #     on_fit_epoch_end: metrikleri toplar ve Excel'e yazar.
    #
    #     on_train_end: log training completion.
    #     on_train_end: egitimin tamamlandigini loglar.

    def __init__(
        self,
        output_path:  Path,
        class_counts: list[int],
        class_names:  list[str],
        beta:         float,
    ):
        # Do NOT create the Excel or its parent directory here.
        # Excel'i veya ust klasorunu burada olusturma.
        # If we create the directory early, YOLO sees it as existing and
        # appends a suffix to the run name (yolov8n_cb_..._2).
        # Eger klasoru erken olusturursak YOLO onu var gorup run ismine
        # son ek ekler (yolov8n_cb_..._2).
        # The Excel is built later in on_pretrain_routine_end when the
        # real YOLO save_dir is known.
        # Excel, YOLO'nun gercek save_dir'i belli olunca
        # on_pretrain_routine_end icinde olusturulur.
        self.output_path  = output_path
        self.class_counts = class_counts
        self.class_names  = class_names
        self.beta         = beta

    def on_pretrain_routine_end(self, trainer) -> None:
        # Inject CB weights after model is on device but before criterion init.
        # Model cihaza tasindiktan sonra fakat criterion init edilmeden once CB agirliklarini inject eder.

        # Unwrap DDP if model is wrapped.
        # Model DDP ile sarilmissa ic modele ulas.
        raw = trainer.model
        while isinstance(raw, nn.parallel.DistributedDataParallel):
            raw = raw.module

        device = next(raw.parameters()).device
        inject_cb_weights(raw, self.class_counts, self.beta, device)

        # Redirect Excel path into the actual YOLO run directory.
        # Excel yolunu gercek YOLO run dizinine yonlendir.
        # This guarantees the Excel lands next to weights/, results.png, etc.
        # Bu sayede Excel weights/, results.png gibi dosyalarla ayni yere duser.
        # If the user passed a filename like "foo/metrics.xlsx", we keep
        # only the file name and drop it inside trainer.save_dir.
        # Kullanici "foo/metrics.xlsx" gibi bir yol verdiyse sadece dosya
        # adini alip trainer.save_dir icine birakiriz.
        save_dir = Path(trainer.save_dir)
        self.output_path = save_dir / self.output_path.name

        if not self.output_path.exists():
            build_excel(self.output_path, self.class_names)
        LOGGER.info(f"[CB-Train] Excel will be written to: {self.output_path}")

    def on_fit_epoch_end(self, trainer) -> None:
        # Collect metrics at the end of an epoch and write them to Excel.
        # Epoch sonunda metrikleri toplar ve Excel'e yazar.
        epoch      = trainer.epoch
        metrics    = trainer.metrics
        loss_items = trainer.loss_items
        lr_list    = trainer.scheduler.get_last_lr()
        lr_val     = lr_list[0] if lr_list else trainer.args.lr0

        # Train loss triplet: box, cls, dfl.
        # Train loss uclusu: box, cls, dfl.
        def _li(i):
            return float(loss_items[i]) if loss_items is not None else None

        box_t, cls_t, dfl_t = _li(0), _li(1), _li(2)

        # Val metric keys depend on the Ultralytics version.
        # Val metrik key'leri Ultralytics surumune gore degisir.
        def _m(key):
            return float(metrics[key]) if key in metrics else None

        box_v    = _m("val/box_loss")
        cls_v    = _m("val/cls_loss")
        dfl_v    = _m("val/dfl_loss")
        prec     = _m("metrics/precision(B)")
        rec      = _m("metrics/recall(B)")
        map50    = _m("metrics/mAP50(B)")
        map50_95 = _m("metrics/mAP50-95(B)")

        f1 = None
        if prec is not None and rec is not None:
            denom = prec + rec
            f1 = (2 * prec * rec / denom) if denom > 0 else 0.0

        # Per-class metrics from the validator.
        # Validator'dan per-class metrikler.
        per_class = self._collect_per_class(trainer)

        # Build row: static values then per-class quadruples.
        # Satir olustur: once static degerler, sonra per-class dortluler.
        row = [
            epoch, round(lr_val, 8),
            _fmt(box_t), _fmt(box_v),
            _fmt(cls_t), _fmt(cls_v),
            _fmt(dfl_t), _fmt(dfl_v),
            _fmt(prec), _fmt(rec), _fmt(f1),
            _fmt(map50), _fmt(map50_95),
        ]
        for i in range(len(self.class_names)):
            row.extend([
                _fmt(per_class["map50"][i]),
                _fmt(per_class["map50_95"][i]),
                _fmt(per_class["precision"][i]),
                _fmt(per_class["recall"][i]),
            ])

        append_epoch_row(self.output_path, row)

        if all(v is not None for v in [box_t, cls_t, dfl_t, map50]):
            LOGGER.info(
                f"[CB-Train] Epoch {epoch:3d} -> "
                f"box={box_t:.4f}  cls={cls_t:.4f}  dfl={dfl_t:.4f}  "
                f"mAP@0.5={map50:.4f}"
            )
        else:
            LOGGER.info(f"[CB-Train] Epoch {epoch} saved.")

    def _collect_per_class(self, trainer) -> dict:
        # Collect per-class metrics from trainer.validator.metrics.
        # trainer.validator.metrics'ten per-class metrikleri toplar.
        #
        # Ultralytics stores per-class metrics in DetMetrics.box object.
        # Ultralytics per-class metrikleri DetMetrics.box nesnesinde tutar.
        nc = len(self.class_names)
        result = {
            "map50":     [None] * nc,
            "map50_95":  [None] * nc,
            "precision": [None] * nc,
            "recall":    [None] * nc,
        }

        validator = getattr(trainer, "validator", None)
        if validator is None:
            return result

        v_metrics = getattr(validator, "metrics", None)
        if v_metrics is None:
            return result

        box = getattr(v_metrics, "box", None)
        if box is None:
            return result

        # Map from class index to position in the metric arrays.
        # Sinif indeksinden metrik dizilerindeki pozisyona haritalama.
        ap_class_index = getattr(box, "ap_class_index", None)
        if ap_class_index is None or len(ap_class_index) == 0:
            return result

        maps     = getattr(box, "maps", None)
        p_arr    = getattr(box, "p", None)
        r_arr    = getattr(box, "recall", None)
        if r_arr is None:
            r_arr = getattr(box, "r", None)
        ap50_arr = getattr(box, "ap50", None)

        for pos, cls_idx in enumerate(ap_class_index):
            cls_idx = int(cls_idx)
            if cls_idx >= nc:
                continue
            if ap50_arr is not None and pos < len(ap50_arr):
                result["map50"][cls_idx] = ap50_arr[pos]
            if maps is not None and cls_idx < len(maps):
                result["map50_95"][cls_idx] = maps[cls_idx]
            if p_arr is not None and pos < len(p_arr):
                result["precision"][cls_idx] = p_arr[pos]
            if r_arr is not None and pos < len(r_arr):
                result["recall"][cls_idx] = r_arr[pos]

        return result

    def on_train_end(self, trainer) -> None:
        LOGGER.info(
            f"[CB-Train] Training finished. Output: {self.output_path}"
        )


# Main training function.
# Ana egitim fonksiyonu.

def train(
    data:         str,
    class_counts: list[int],
    cfg:          Optional[dict] = None,
    class_names:  Optional[list[str]] = None,
) -> None:
    # Start YOLOv8-nano training.
    # YOLOv8-nano egitimini baslatir.
    #
    # Uses NANO_CFG if cfg is None, otherwise overrides only the given keys.
    # cfg None ise NANO_CFG kullanilir, aksi halde sadece verilen anahtarlar override edilir.
    #
    # If class_names is None, names are read from data.yaml.
    # class_names None ise isimler data.yaml'dan okunur.
    c = {**NANO_CFG, **(cfg or {})}

    # Resolve class names.
    # Sinif isimlerini belirle.
    if class_names is None:
        class_names = read_class_names(data)

    # Consistency check between class_counts and class_names.
    # class_counts ile class_names arasinda tutarlilik kontrolu.
    if len(class_counts) != len(class_names):
        raise ValueError(
            f"class_counts length ({len(class_counts)}) != "
            f"class_names length ({len(class_names)}). "
            f"data.yaml: {class_names}, counts: {class_counts}"
        )

    # Resolve Excel output path relative to runs/ if not absolute.
    # Mutlak degilse Excel cikti yolunu runs/ dizinine gore coz.
    excel_path = resolve_excel_path(c["OUTPUT_XLSX"])

    LOGGER.info(
        f"[CB-Train] Classes: {class_names}, "
        f"counts: {class_counts}, seed: {c['SEED']}"
    )
    LOGGER.info(f"[CB-Train] Runs directory: {c['PROJECT']}")
    LOGGER.info(f"[CB-Train] Excel output: {excel_path}")

    model = YOLO("yolov8n.pt")

    cbs = CBTrainingCallbacks(
        output_path=excel_path,
        class_counts=class_counts,
        class_names=class_names,
        beta=c["BETA"],
    )

    model.add_callback("on_pretrain_routine_end", cbs.on_pretrain_routine_end)
    model.add_callback("on_fit_epoch_end",        cbs.on_fit_epoch_end)
    model.add_callback("on_train_end",            cbs.on_train_end)

    model.train(
        data=data,
        imgsz=c["IMG_SIZE"],
        device=c["DEVICE"],
        epochs=c["EPOCHS"],
        batch=c["BATCH_SIZE"],
        patience=c["PATIENCE"],
        workers=c["WORKERS"],
        cache=c["CACHE"],
        seed=c["SEED"],
        optimizer=c["OPTIMIZER"],
        lr0=c["LR0"],
        lrf=c["LRF"],
        momentum=c["MOMENTUM"],
        weight_decay=c["WEIGHT_DECAY"],
        warmup_epochs=c["WARMUP_EPOCHS"],
        amp=c["AMP"],
        hsv_h=c["HSV_H"],
        hsv_s=c["HSV_S"],
        hsv_v=c["HSV_V"],
        mosaic=c["MOSAIC"],
        mixup=c["MIXUP"],
        flipud=c["FLIPUD"],
        fliplr=c["FLIPLR"],
        degrees=c["DEGREES"],
        translate=c["TRANSLATE"],
        scale=c["SCALE"],
        perspective=c["PERSPECTIVE"],
        erasing=c["ERASING"],
        project=c["PROJECT"],
        name=c["NAME"],
        verbose=True,
    )


# CLI entry point.
# CLI giris noktasi.

def parse_args():
    p = argparse.ArgumentParser(
        description="YOLOv8-nano + Class-Balanced Loss training",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )

    p.add_argument("--data",   required=True,
                   help="Path to data.yaml / data.yaml yolu")
    p.add_argument("--counts", type=int, nargs="+", required=True,
                   help="Sample count per class / Sinif basina ornek sayisi")

    p.add_argument("--imgsz",         type=int,   default=NANO_CFG["IMG_SIZE"])
    p.add_argument("--device",                    default=str(NANO_CFG["DEVICE"]))
    p.add_argument("--epochs",        type=int,   default=NANO_CFG["EPOCHS"])
    p.add_argument("--batch",         type=int,   default=NANO_CFG["BATCH_SIZE"])
    p.add_argument("--patience",      type=int,   default=NANO_CFG["PATIENCE"])
    p.add_argument("--workers",       type=int,   default=NANO_CFG["WORKERS"])
    p.add_argument("--cache",                     default=str(NANO_CFG["CACHE"]))
    p.add_argument("--seed",          type=int,   default=NANO_CFG["SEED"],
                   help="Random seed for reproducibility / Tekrarlanabilirlik icin seed")

    p.add_argument("--optimizer",                 default=NANO_CFG["OPTIMIZER"])
    p.add_argument("--lr0",           type=float, default=NANO_CFG["LR0"])
    p.add_argument("--lrf",           type=float, default=NANO_CFG["LRF"])
    p.add_argument("--momentum",      type=float, default=NANO_CFG["MOMENTUM"])
    p.add_argument("--weight-decay",  type=float, default=NANO_CFG["WEIGHT_DECAY"],
                   dest="weight_decay")
    p.add_argument("--warmup-epochs", type=int,   default=NANO_CFG["WARMUP_EPOCHS"],
                   dest="warmup_epochs")
    p.add_argument("--no-amp",        action="store_false", dest="amp")
    p.set_defaults(amp=NANO_CFG["AMP"])

    p.add_argument("--hsv-h",       type=float, default=NANO_CFG["HSV_H"],       dest="hsv_h")
    p.add_argument("--hsv-s",       type=float, default=NANO_CFG["HSV_S"],       dest="hsv_s")
    p.add_argument("--hsv-v",       type=float, default=NANO_CFG["HSV_V"],       dest="hsv_v")
    p.add_argument("--mosaic",      type=float, default=NANO_CFG["MOSAIC"])
    p.add_argument("--mixup",       type=float, default=NANO_CFG["MIXUP"])
    p.add_argument("--flipud",      type=float, default=NANO_CFG["FLIPUD"])
    p.add_argument("--fliplr",      type=float, default=NANO_CFG["FLIPLR"])
    p.add_argument("--degrees",     type=float, default=NANO_CFG["DEGREES"])
    p.add_argument("--translate",   type=float, default=NANO_CFG["TRANSLATE"])
    p.add_argument("--scale",       type=float, default=NANO_CFG["SCALE"])
    p.add_argument("--perspective", type=float, default=NANO_CFG["PERSPECTIVE"])
    p.add_argument("--erasing",     type=float, default=NANO_CFG["ERASING"])

    p.add_argument("--beta",    type=float, default=NANO_CFG["BETA"],
                   help="CB Loss beta (0 < beta < 1)")
    p.add_argument("--project",             default=NANO_CFG["PROJECT"])
    p.add_argument("--name",                default=NANO_CFG["NAME"])
    p.add_argument("--output",              default=NANO_CFG["OUTPUT_XLSX"],
                   help="Excel output file / Excel cikti dosyasi")

    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()

    override = {
        "IMG_SIZE":      args.imgsz,
        "DEVICE":        args.device,
        "EPOCHS":        args.epochs,
        "BATCH_SIZE":    args.batch,
        "PATIENCE":      args.patience,
        "WORKERS":       args.workers,
        "CACHE":         args.cache,
        "SEED":          args.seed,
        "OPTIMIZER":     args.optimizer,
        "LR0":           args.lr0,
        "LRF":           args.lrf,
        "MOMENTUM":      args.momentum,
        "WEIGHT_DECAY":  args.weight_decay,
        "WARMUP_EPOCHS": args.warmup_epochs,
        "AMP":           args.amp,
        "HSV_H":         args.hsv_h,
        "HSV_S":         args.hsv_s,
        "HSV_V":         args.hsv_v,
        "MOSAIC":        args.mosaic,
        "MIXUP":         args.mixup,
        "FLIPUD":        args.flipud,
        "FLIPLR":        args.fliplr,
        "DEGREES":       args.degrees,
        "TRANSLATE":     args.translate,
        "SCALE":         args.scale,
        "PERSPECTIVE":   args.perspective,
        "ERASING":       args.erasing,
        "BETA":          args.beta,
        "PROJECT":       args.project,
        "NAME":          args.name,
        "OUTPUT_XLSX":   args.output,
    }

    train(
        data=args.data,
        class_counts=args.counts,
        cfg=override,
    )