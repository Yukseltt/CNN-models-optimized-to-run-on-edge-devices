# train_unet.py
#
# Training script for U-Net semantic segmentation on thermal data.
# Termal veri uzerinde U-Net semantik segmentasyon egitim scripti.
#
# Same training-loop conventions as effecientnetB0/train_efficientnet_b0.py:
# manual warmup + cosine LR, per-epoch validation, Excel logging, best/last
# checkpointing, early stopping, resume from last.pt.
# effecientnetB0/train_efficientnet_b0.py ile ayni egitim kuralllari:
# manuel warmup + cosine LR, her epoch validation, Excel logging, best/last
# checkpoint, early stopping, last.pt'den resume.

import math
import random
import time
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from torch.amp import GradScaler, autocast
from torch.optim import SGD
from torch.utils.data import DataLoader

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))
from dataset import UNetSegDataset, compute_class_pixel_counts
from model_factory import create_unet, count_parameters


# Path constants / Yol sabitleri.

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
RUNS_DIR     = PROJECT_ROOT / "runs"


# Default config / Varsayilan konfigurasyon.

DEFAULT_CFG = {
    "EPOCHS":         50,
    "BATCH_SIZE":     64,
    "PATIENCE":       10,
    "WORKERS":        4,
    "SEED":           0,
    "DEVICE":         "cuda:0",

    # Manuel warmup + cosine.
    "LR0":            0.01,
    "LRF":            0.01,
    "MOMENTUM":       0.9,
    "WEIGHT_DECAY":   0.0001,
    "NESTEROV":       True,
    "WARMUP_EPOCHS":  2,

    # Encoder fine-tuning multiplier'i.
    "ENCODER_LR_MULT": 0.1,
    "DECODER_LR_MULT": 1.0,

    # Model.
    "ENCODER_NAME":    "resnet34",
    "ENCODER_WEIGHTS": "imagenet",
    "PRETRAINED_PATH": None,
    "NUM_CLASSES":     4,           # bg + Person + Car + OtherVehicle
    "INPUT_SIZE":      512,
    "USE_AMP":         True,

    # Sinif dengesizligi icin opsiyonlar:
    #   "USE_CLASS_WEIGHTS": True ise train setinden hesaplanan
    #   inverse-frequency agirliklari CrossEntropy'ye verilir.
    "USE_CLASS_WEIGHTS": True,
    "IGNORE_INDEX":      -100,

    "PROJECT":     str(RUNS_DIR),
    "NAME":        "unet_default",
    "OUTPUT_XLSX": "training_metrics.xlsx",

    "LOG_INTERVAL": 50,

    # Bir checkpoint yoluna ayarlanirsa, oradan devam eder.
    "RESUME_FROM": None,
}


# Helpers / Yardimcilar.

def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def compute_lr(
    epoch:         int,
    total_epochs:  int,
    warmup_epochs: int,
    lr0:           float,
    lrf:           float,
) -> float:
    if epoch < warmup_epochs:
        return lr0 * (epoch + 1) / max(1, warmup_epochs)
    progress = (epoch - warmup_epochs) / max(1, total_epochs - warmup_epochs)
    cosine   = 0.5 * (1 + math.cos(math.pi * progress))
    return lr0 * (lrf + (1 - lrf) * cosine)


def resolve_run_dir(project_dir: str, base_name: str) -> Path:
    project = Path(project_dir)
    project.mkdir(parents=True, exist_ok=True)

    if not (project / base_name).exists():
        return project / base_name

    counter = 2
    while (project / f"{base_name}{counter}").exists():
        counter += 1
    return project / f"{base_name}{counter}"


# Dataloaders / Dataloader'lar.

def build_dataloaders(data_dir: str, cfg: dict) -> Tuple[DataLoader, DataLoader, UNetSegDataset]:
    data_dir = Path(data_dir)

    train_ds = UNetSegDataset(
        images_dir=str(data_dir / "train" / "images"),
        masks_dir=str(data_dir / "train" / "masks"),
        input_size=cfg["INPUT_SIZE"],
        augment=True,
    )
    val_ds = UNetSegDataset(
        images_dir=str(data_dir / "val" / "images"),
        masks_dir=str(data_dir / "val" / "masks"),
        input_size=cfg["INPUT_SIZE"],
        augment=False,
    )

    train_loader = DataLoader(
        train_ds,
        batch_size=cfg["BATCH_SIZE"],
        shuffle=True,
        num_workers=cfg["WORKERS"],
        pin_memory=True,
        drop_last=True,
        persistent_workers=(cfg["WORKERS"] > 0),
    )
    val_loader = DataLoader(
        val_ds,
        batch_size=cfg["BATCH_SIZE"],
        shuffle=False,
        num_workers=cfg["WORKERS"],
        pin_memory=True,
        drop_last=False,
        persistent_workers=(cfg["WORKERS"] > 0),
    )
    return train_loader, val_loader, train_ds


# Optimizer / Optimizer.

def build_optimizer(model: nn.Module, cfg: dict) -> torch.optim.Optimizer:
    # smp.Unet -> model.encoder + model.decoder + model.segmentation_head
    encoder_params = list(model.encoder.parameters())
    decoder_params = list(model.decoder.parameters()) + list(
        model.segmentation_head.parameters()
    )
    return SGD(
        [
            {
                "params": encoder_params,
                "lr":     cfg["LR0"] * cfg["ENCODER_LR_MULT"],
                "name":   "encoder",
            },
            {
                "params": decoder_params,
                "lr":     cfg["LR0"] * cfg["DECODER_LR_MULT"],
                "name":   "decoder",
            },
        ],
        momentum=cfg["MOMENTUM"],
        weight_decay=cfg["WEIGHT_DECAY"],
        nesterov=cfg["NESTEROV"],
    )


def set_param_group_lrs(optimizer: torch.optim.Optimizer, base_lr: float, cfg: dict) -> None:
    mults = {
        "encoder": cfg["ENCODER_LR_MULT"],
        "decoder": cfg["DECODER_LR_MULT"],
    }
    for g in optimizer.param_groups:
        g["lr"] = base_lr * mults[g["name"]]


# Excel logger / Excel logger.

HEADERS = [
    "Epoch", "Learning Rate",
    "Train Loss", "Train mIoU",  "Train Pixel Acc",  "Train Mean Dice",
    "Val Loss",   "Val mIoU",    "Val Pixel Acc",    "Val Mean Dice",
]

_THIN       = Side(border_style="thin", color="BBBBBB")
BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)


def build_excel(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Training Metrics"
    for ci, val in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"
    wb.save(path)


def append_epoch_row(path: Path, values: list) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb["Training Metrics"]
    r  = ws.max_row + 1
    alt = PatternFill("solid", start_color="EBF3FB") if r % 2 == 0 else None
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font      = DATA_FONT
        c.alignment = Alignment(horizontal="center")
        c.border    = BORDER
        if alt:
            c.fill = alt
    wb.save(path)


def _fmt(v):
    return None if v is None else round(float(v), 4)


# Metrics / Metrikler.

class SegConfusion:
    # Confusion matrix-tabanli segmentasyon metrikleri (mIoU, Dice, Pixel Acc).

    def __init__(self, num_classes: int):
        self.num_classes = num_classes
        self.mat = torch.zeros((num_classes, num_classes), dtype=torch.int64)

    def update(self, preds: torch.Tensor, targets: torch.Tensor) -> None:
        # preds, targets: (N, H, W) long tensors
        n = self.num_classes
        valid = (targets >= 0) & (targets < n)
        p = preds[valid].view(-1)
        t = targets[valid].view(-1)
        idx = t * n + p
        bc  = torch.bincount(idx, minlength=n * n).view(n, n)
        self.mat += bc.to(self.mat.device)

    def compute(self) -> dict:
        m   = self.mat.double()
        tp  = m.diag()
        gt  = m.sum(dim=1)
        pr  = m.sum(dim=0)
        union = gt + pr - tp

        iou  = torch.where(union > 0, tp / union.clamp(min=1), torch.full_like(tp, float("nan")))
        dice = torch.where((gt + pr) > 0, 2 * tp / (gt + pr).clamp(min=1), torch.full_like(tp, float("nan")))

        miou      = torch.nanmean(iou).item()
        mean_dice = torch.nanmean(dice).item()
        pix_acc   = (tp.sum() / m.sum().clamp(min=1)).item()

        return {
            "miou":      miou,
            "mean_dice": mean_dice,
            "pixel_acc": pix_acc,
            "iou_per_class":  iou.tolist(),
            "dice_per_class": dice.tolist(),
        }


# Class weights / Sinif agirliklari.

def compute_inverse_freq_weights(
    train_ds:    UNetSegDataset,
    num_classes: int,
) -> torch.Tensor:
    print("[UNET] Computing class pixel counts on train set...")
    counts = compute_class_pixel_counts(train_ds, num_classes).astype(np.float64)
    counts = np.maximum(counts, 1.0)
    freq   = counts / counts.sum()
    inv    = 1.0 / freq
    weights = inv / inv.mean()
    print(f"[UNET] Class pixel counts: {counts.astype(np.int64).tolist()}")
    print(f"[UNET] Class weights:      {[round(w, 4) for w in weights.tolist()]}")
    return torch.tensor(weights, dtype=torch.float32)


# Train epoch / Train epoch.

def train_one_epoch(
    model:        nn.Module,
    optimizer:    torch.optim.Optimizer,
    criterion:    nn.Module,
    loader:       DataLoader,
    device:       str,
    epoch:        int,
    scaler:       Optional[GradScaler],
    use_amp:      bool,
    log_interval: int,
    num_classes:  int,
) -> dict:
    model.train()

    loss_sum  = 0.0
    n_samples = 0

    conf = SegConfusion(num_classes)

    epoch_start = time.time()

    for batch_idx, (images, masks) in enumerate(loader):
        images = images.to(device, non_blocking=True)
        masks  = masks.to(device,  non_blocking=True)

        optimizer.zero_grad(set_to_none=True)

        if use_amp:
            with autocast("cuda"):
                logits = model(images)
                loss   = criterion(logits, masks)
            scaler.scale(loss).backward()
            scaler.step(optimizer)
            scaler.update()
        else:
            logits = model(images)
            loss   = criterion(logits, masks)
            loss.backward()
            optimizer.step()

        with torch.no_grad():
            preds = logits.argmax(dim=1)
            conf.update(preds.detach().cpu(), masks.detach().cpu())

        bs         = images.size(0)
        loss_sum  += loss.item() * bs
        n_samples += bs

        if (batch_idx + 1) % log_interval == 0:
            elapsed  = time.time() - epoch_start
            avg_loss = loss_sum / max(1, n_samples)
            print(
                f"  [Epoch {epoch}] batch {batch_idx + 1}/{len(loader)}  "
                f"avg_loss={avg_loss:.4f}  elapsed={elapsed:.1f}s"
            )

    metrics = conf.compute()
    return {
        "loss":      loss_sum / max(1, n_samples),
        "miou":      metrics["miou"],
        "pixel_acc": metrics["pixel_acc"],
        "mean_dice": metrics["mean_dice"],
        "iou_per_class": metrics["iou_per_class"],
    }


# Validation / Validation.

@torch.no_grad()
def validate(
    model:       nn.Module,
    criterion:   nn.Module,
    loader:      DataLoader,
    device:      str,
    num_classes: int,
) -> dict:
    model.eval()

    loss_sum  = 0.0
    n_samples = 0
    conf = SegConfusion(num_classes)

    val_start = time.time()

    for images, masks in loader:
        images = images.to(device, non_blocking=True)
        masks  = masks.to(device,  non_blocking=True)

        logits = model(images)
        loss   = criterion(logits, masks)
        preds  = logits.argmax(dim=1)

        bs         = images.size(0)
        loss_sum  += loss.item() * bs
        n_samples += bs
        conf.update(preds.cpu(), masks.cpu())

    metrics = conf.compute()
    elapsed = time.time() - val_start
    print(f"  Validation done in {elapsed:.1f}s")

    return {
        "loss":      loss_sum / max(1, n_samples),
        "miou":      metrics["miou"],
        "pixel_acc": metrics["pixel_acc"],
        "mean_dice": metrics["mean_dice"],
        "iou_per_class": metrics["iou_per_class"],
    }


# Resume / Resume.

def load_resume_state(
    resume_path: str,
    model:       nn.Module,
    optimizer:   torch.optim.Optimizer,
    device:      str,
) -> dict:
    print(f"[UNET] Resuming from {resume_path}")
    ckpt = torch.load(resume_path, map_location=device, weights_only=False)

    model.load_state_dict(ckpt["model"])
    optimizer.load_state_dict(ckpt["optimizer"])

    start_epoch = int(ckpt["epoch"]) + 1
    print(f"[UNET] Resumed at epoch {start_epoch} "
          f"(checkpoint saved at epoch {ckpt['epoch']}).")
    return {"start_epoch": start_epoch}


def load_best_state(best_path: Path) -> dict:
    if not best_path.exists():
        return {"best_metric": -1.0, "best_epoch": -1}
    bckpt = torch.load(best_path, map_location="cpu", weights_only=False)
    return {
        "best_metric": float(bckpt.get("metric", -1.0)),
        "best_epoch":  int(bckpt.get("epoch", -1)),
    }


# Main / Main.

def train(data_dir: str, cfg: Optional[dict] = None) -> None:
    c = {**DEFAULT_CFG, **(cfg or {})}

    set_seed(c["SEED"])

    if c["DEVICE"].startswith("cuda") and not torch.cuda.is_available():
        print("[UNET] CUDA not available, falling back to CPU.")
        c["DEVICE"] = "cpu"
        c["USE_AMP"] = False

    is_resume = c.get("RESUME_FROM") is not None

    print(f"[UNET] Model:    U-Net ({c['ENCODER_NAME']} encoder)")
    print(f"[UNET] Device:   {c['DEVICE']}")
    print(f"[UNET] Batch:    {c['BATCH_SIZE']}, Epochs: {c['EPOCHS']}")
    print(f"[UNET] LR0:      {c['LR0']}, LRF: {c['LRF']}")
    print(f"[UNET] Input:    {c['INPUT_SIZE']}x{c['INPUT_SIZE']}")
    print(f"[UNET] Classes:  {c['NUM_CLASSES']}")
    print(f"[UNET] AMP:      {c['USE_AMP']}")
    print(f"[UNET] Patience: {c['PATIENCE']}")
    if is_resume:
        print(f"[UNET] Resume:   {c['RESUME_FROM']}")

    train_loader, val_loader, train_ds = build_dataloaders(data_dir, c)
    print(f"[UNET] Train: {len(train_loader.dataset)} images, {len(train_loader)} batches")
    print(f"[UNET] Val:   {len(val_loader.dataset)} images, {len(val_loader)} batches")

    model = create_unet(
        num_classes=c["NUM_CLASSES"],
        encoder_name=c["ENCODER_NAME"],
        encoder_weights=c["ENCODER_WEIGHTS"] if not is_resume else None,
        pretrained_path=c["PRETRAINED_PATH"],
    ).to(c["DEVICE"])

    params = count_parameters(model)
    print(f"[UNET] Total params: {params['total_millions']:.2f}M")

    # Loss.
    if c["USE_CLASS_WEIGHTS"]:
        weights = compute_inverse_freq_weights(train_ds, c["NUM_CLASSES"]).to(c["DEVICE"])
        criterion = nn.CrossEntropyLoss(weight=weights, ignore_index=c["IGNORE_INDEX"])
    else:
        criterion = nn.CrossEntropyLoss(ignore_index=c["IGNORE_INDEX"])

    optimizer = build_optimizer(model, c)

    use_amp = c["USE_AMP"] and c["DEVICE"].startswith("cuda")
    scaler  = GradScaler("cuda") if use_amp else None
    if use_amp:
        torch.backends.cudnn.benchmark = True

    start_epoch      = 0
    best_metric      = -1.0
    best_epoch       = -1
    patience_counter = 0

    if is_resume:
        save_dir   = Path(c["RESUME_FROM"]).resolve().parent
        excel_path = save_dir / c["OUTPUT_XLSX"]
        if not excel_path.exists():
            raise FileNotFoundError(
                f"Excel not found in resume dir: {excel_path}. "
                f"Cannot resume without existing metrics file."
            )
        print(f"[UNET] Resume dir: {save_dir}")

        resume_state = load_resume_state(c["RESUME_FROM"], model, optimizer, c["DEVICE"])
        start_epoch  = resume_state["start_epoch"]

        best_state       = load_best_state(save_dir / "best.pt")
        best_metric      = best_state["best_metric"]
        best_epoch       = best_state["best_epoch"]
        patience_counter = max(0, (start_epoch - 1) - best_epoch)
        print(f"[UNET] Best so far: val_mIoU={best_metric:.4f} @ epoch {best_epoch}")
        print(f"[UNET] Patience counter: {patience_counter}/{c['PATIENCE']}")
    else:
        save_dir = resolve_run_dir(c["PROJECT"], c["NAME"])
        save_dir.mkdir(parents=True, exist_ok=True)
        excel_path = save_dir / c["OUTPUT_XLSX"]
        print(f"[UNET] Output dir: {save_dir}")
        build_excel(excel_path)

    print(f"\n[UNET] Starting training from epoch {start_epoch}...\n")

    for epoch in range(start_epoch, c["EPOCHS"]):
        epoch_start = time.time()

        base_lr = compute_lr(
            epoch=epoch,
            total_epochs=c["EPOCHS"],
            warmup_epochs=c["WARMUP_EPOCHS"],
            lr0=c["LR0"],
            lrf=c["LRF"],
        )
        set_param_group_lrs(optimizer, base_lr, c)

        print(f"=== Epoch {epoch}/{c['EPOCHS'] - 1}  base_lr={base_lr:.6f} ===")

        train_metrics = train_one_epoch(
            model=model,
            optimizer=optimizer,
            criterion=criterion,
            loader=train_loader,
            device=c["DEVICE"],
            epoch=epoch,
            scaler=scaler,
            use_amp=use_amp,
            log_interval=c["LOG_INTERVAL"],
            num_classes=c["NUM_CLASSES"],
        )
        print(f"  Train: loss={train_metrics['loss']:.4f}  "
              f"mIoU={train_metrics['miou']:.4f}  "
              f"pix_acc={train_metrics['pixel_acc']:.4f}  "
              f"dice={train_metrics['mean_dice']:.4f}")

        val_metrics = validate(
            model=model,
            criterion=criterion,
            loader=val_loader,
            device=c["DEVICE"],
            num_classes=c["NUM_CLASSES"],
        )
        print(f"  Val:   loss={val_metrics['loss']:.4f}  "
              f"mIoU={val_metrics['miou']:.4f}  "
              f"pix_acc={val_metrics['pixel_acc']:.4f}  "
              f"dice={val_metrics['mean_dice']:.4f}")
        print(f"  Val IoU per class: "
              f"{[round(x, 4) if not (x != x) else 'nan' for x in val_metrics['iou_per_class']]}")

        append_epoch_row(
            excel_path,
            [
                epoch,
                round(base_lr, 8),
                _fmt(train_metrics["loss"]),
                _fmt(train_metrics["miou"]),
                _fmt(train_metrics["pixel_acc"]),
                _fmt(train_metrics["mean_dice"]),
                _fmt(val_metrics["loss"]),
                _fmt(val_metrics["miou"]),
                _fmt(val_metrics["pixel_acc"]),
                _fmt(val_metrics["mean_dice"]),
            ],
        )

        torch.save(
            {
                "epoch":     epoch,
                "model":     model.state_dict(),
                "optimizer": optimizer.state_dict(),
                "cfg":       c,
            },
            save_dir / "last.pt",
        )

        # Best-model selection: val_mIoU.
        if val_metrics["miou"] > best_metric:
            best_metric      = val_metrics["miou"]
            best_epoch       = epoch
            patience_counter = 0
            torch.save(
                {
                    "epoch":     epoch,
                    "model":     model.state_dict(),
                    "optimizer": optimizer.state_dict(),
                    "cfg":       c,
                    "metric":    val_metrics["miou"],
                },
                save_dir / "best.pt",
            )
            print(f"  >> NEW BEST val_mIoU={best_metric:.4f} at epoch {best_epoch}")
        else:
            patience_counter += 1
            print(f"  No improvement ({patience_counter}/{c['PATIENCE']})  "
                  f"best val_mIoU={best_metric:.4f} @ epoch {best_epoch}")

        epoch_time = time.time() - epoch_start
        print(f"  Epoch time: {epoch_time:.1f}s\n")

        if patience_counter >= c["PATIENCE"]:
            print(f"[UNET] Early stopping at epoch {epoch}. "
                  f"Best: epoch {best_epoch}, val_mIoU={best_metric:.4f}")
            break

    print(f"\n[UNET] Training finished.")
    print(f"  Best epoch:    {best_epoch}")
    print(f"  Best val_mIoU: {best_metric:.4f}")
    print(f"  Output dir:    {save_dir}")
