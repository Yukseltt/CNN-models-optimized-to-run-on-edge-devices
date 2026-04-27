"""
epoch_metrics_logger.py
========================
HuggingFace Trainer için epoch-bazlı metrik logger'ı.

Her epoch sonunda bir satır, `yolov8/training_results.xlsx` ile aynı
iki seviyeli başlık şeması kullanılarak .xlsx'e yazılır:

    Epoch | Learning Rate | Box Loss (Train/Val) | Class Loss (Train/Val)
          | DFL Loss (Train/Val) | Precision (Val/Train*) | Recall (Val/Train*)
          | F1 Score (Val/Train*) | mAP (mAP@0.5 / mAP@0.5:0.95) | Accuracy (val)

Bir sütunun karşılığı gelen log'da yoksa (örn. RT-DETR'de DFL yoktur,
varsayılan mAP evaluator'ı Precision/Recall döndürmez) o hücre boş kalır.

KULLANIM
--------
    from epoch_metrics_logger import EpochMetricsXlsxCallback

    metrics_cb = EpochMetricsXlsxCallback("training_results.xlsx")

    trainer = Trainer(
        model=model, args=training_args,
        train_dataset=train_dataset, eval_dataset=validation_dataset,
        processing_class=image_processor, data_collator=collate_fn,
        compute_metrics=eval_compute_metrics_fn,
        callbacks=[metrics_cb],
    )
"""

import math
from pathlib import Path

from openpyxl import Workbook
from transformers import TrainerCallback


# ──────────────────────────────────────────────────────────────────────
# Sütun şeması — yolov8/training_results.xlsx ile birebir aynı
# ──────────────────────────────────────────────────────────────────────
COLUMNS = [
    ("Epoch",          ""),
    ("Learning Rate",  ""),
    ("Box Loss",       "Train"),
    ("Box Loss",       "Val"),
    ("Class Loss",     "Train"),
    ("Class Loss",     "Val"),
    ("DFL Loss",       "Train"),
    ("DFL Loss",       "Val"),
    ("Precision",      "Val"),
    ("Precision",      "Train*"),
    ("Recall",         "Val"),
    ("Recall",         "Train*"),
    ("F1 Score",       "Val"),
    ("F1 Score",       "Train*"),
    ("mAP",            "mAP@0.5"),
    ("mAP",            "mAP@0.5:0.95"),
    ("Accuracy (val)", ""),
]


def _is_num(v):
    return v is not None and not (isinstance(v, float) and math.isnan(v))


def _get(d, *keys):
    for k in keys:
        v = d.get(k)
        if _is_num(v):
            return v
    return None


def _f1(p, r):
    if p is None or r is None:
        return None
    s = p + r
    return 0.0 if s == 0 else 2 * p * r / s


class EpochMetricsXlsxCallback(TrainerCallback):
    """
    HF Trainer'ı izler; her epoch için bir satır olacak şekilde
    metrikleri .xlsx'e yazar.

    Parametreler
    ------------
    output_path : str | Path
        Yazılacak xlsx dosyasının yolu.
    loss_key_map : dict, optional
        Gerçek log anahtarlarını bu şemaya eşler. Örn. RT-DETR
        bileşen loss'larını loglarsan:
            {
                "box_train": "loss_bbox",
                "cls_train": "loss_vfl",
                "box_val":   "eval_loss_bbox",
                "cls_val":   "eval_loss_vfl",
            }
        Verilmezse yaygın isimler denenir; bulunamazsa hücre boş kalır.
    """

    def __init__(self, output_path, loss_key_map=None):
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.loss_key_map = loss_key_map or {}

        # epoch -> row dict (son evaluate sonucu bir önceki yazımı günceller)
        self._rows_by_epoch = {}
        # son training-log içeriği
        self._train_buf = {}
        self._lr = None

    # ── HF TrainerCallback kancaları ─────────────────────────────
    def on_log(self, args, state, control, logs=None, **kwargs):
        if not logs:
            return
        # Evaluate log'unu burada tutmuyoruz — on_evaluate'te zaten
        # `metrics` parametresi üzerinden alıyoruz.
        if any(k.startswith("eval_") for k in logs):
            return
        self._train_buf = dict(logs)
        if "learning_rate" in logs:
            self._lr = logs["learning_rate"]

    def on_epoch_end(self, args, state, control, **kwargs):
        epoch = int(round(state.epoch))
        # eval_strategy="no" durumunda bile en az train satırını yaz
        if epoch not in self._rows_by_epoch:
            self._rows_by_epoch[epoch] = self._build_row(epoch, {})
            self._flush()

    def on_evaluate(self, args, state, control, metrics=None, **kwargs):
        epoch = int(round(state.epoch))
        self._rows_by_epoch[epoch] = self._build_row(epoch, metrics or {})
        self._flush()

    # ── Satır oluşturma ─────────────────────────────────────────
    def _build_row(self, epoch, eval_metrics):
        t = self._train_buf
        e = eval_metrics or {}
        m = self.loss_key_map

        box_train = _get(t, m.get("box_train", "loss_bbox"), "loss_bbox", "train_loss_bbox")
        cls_train = _get(t, m.get("cls_train", "loss_vfl"),  "loss_vfl",  "loss_ce", "train_loss_vfl")
        dfl_train = _get(t, m.get("dfl_train", "loss_dfl"),  "loss_dfl")

        box_val = _get(e, m.get("box_val", "eval_loss_bbox"))
        cls_val = _get(e, m.get("cls_val", "eval_loss_vfl"), "eval_loss_ce")
        dfl_val = _get(e, m.get("dfl_val", "eval_loss_dfl"))

        map50   = _get(e, "eval_map_50")
        map5095 = _get(e, "eval_map")

        prec_val = _get(e, "eval_precision")
        rec_val  = _get(e, "eval_recall")
        f1_val   = _f1(prec_val, rec_val)

        acc_val  = _get(e, "eval_accuracy")

        return {
            ("Epoch",          ""):             epoch,
            ("Learning Rate",  ""):             self._lr,
            ("Box Loss",       "Train"):        box_train,
            ("Box Loss",       "Val"):          box_val,
            ("Class Loss",     "Train"):        cls_train,
            ("Class Loss",     "Val"):          cls_val,
            ("DFL Loss",       "Train"):        dfl_train,
            ("DFL Loss",       "Val"):          dfl_val,
            ("Precision",      "Val"):          prec_val,
            ("Precision",      "Train*"):       None,
            ("Recall",         "Val"):          rec_val,
            ("Recall",         "Train*"):       None,
            ("F1 Score",       "Val"):          f1_val,
            ("F1 Score",       "Train*"):       None,
            ("mAP",            "mAP@0.5"):      map50,
            ("mAP",            "mAP@0.5:0.95"): map5095,
            ("Accuracy (val)", ""):             acc_val,
        }

    def _flush(self):
        if not self._rows_by_epoch:
            return

        upper = [c[0] for c in COLUMNS]
        lower = [c[1] for c in COLUMNS]

        wb = Workbook()
        ws = wb.active
        ws.append(upper)
        ws.append(lower)

        # Üst başlık: yatay merge (aynı üst başlık grupları)
        # Alt başlığı boş olan kolonlar için iki satır dikey merge
        i = 0
        n = len(COLUMNS)
        while i < n:
            j = i
            while j + 1 < n and upper[j + 1] == upper[i]:
                j += 1
            if j > i:
                ws.merge_cells(start_row=1, start_column=i + 1,
                               end_row=1, end_column=j + 1)
            if j == i and lower[i] == "":
                ws.merge_cells(start_row=1, start_column=i + 1,
                               end_row=2, end_column=i + 1)
            i = j + 1

        for key in sorted(self._rows_by_epoch):
            row = self._rows_by_epoch[key]
            ws.append([row[c] for c in COLUMNS])

        wb.save(self.output_path)
