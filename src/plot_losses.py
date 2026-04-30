# plot_losses.py
#
# Training metric plot generator.
# Egitim metrigi grafik ureticisi.
#
# Supports three input formats / Uc giris formatini destekler:
#     excel        : Our standard YOLO Excel format (default)
#     excel        : Bizim standart YOLO Excel formatimiz (varsayilan)
#     yolo         : Ultralytics results.csv
#     yolo         : Ultralytics results.csv
#     faster_rcnn  : Faster R-CNN Excel format (4 loss components, val losses)
#     faster_rcnn  : Faster R-CNN Excel formati (4 loss component, val loss'lari)
#
# Usage / Kullanim:
#     python plot_losses.py --input training_results.xlsx --out ./plots
#     python plot_losses.py --input training_results.xlsx --per-class
#     python plot_losses.py --input results.csv --format yolo
#     python plot_losses.py --input training_metrics.xlsx --format faster_rcnn --per-class

import argparse
import csv
import io
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec


# Style constants.
# Stil sabitleri.

TRAIN_COLOR = "#2166AC"
VAL_COLOR   = "#D73027"
CLS_COLOR   = "#D73027"
DFL_COLOR   = "#1A9850"
LW          = 2.0

CLASS_COLORS = [
    "#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
    "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF",
]

FRCNN_COLORS = {
    "classifier":  "#D73027",
    "box_reg":     "#2166AC",
    "objectness":  "#1A9850",
    "rpn_box_reg": "#9467BD",
}

plt.rcParams.update({
    "figure.facecolor":  "white",
    "axes.facecolor":    "white",
    "axes.grid":         True,
    "grid.color":        "#e0e0e0",
    "grid.linewidth":    0.8,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "font.family":       "DejaVu Sans",
    "font.size":         11,
    "axes.titlesize":    13,
    "axes.titleweight":  "bold",
    "axes.labelsize":    11,
    "legend.fontsize":   10,
    "legend.framealpha": 0.9,
    "legend.edgecolor":  "#cccccc",
})


# Data loaders.
# Veri yukleyiciler.

def _to_float_array(values) -> np.ndarray:
    out = []
    for v in values:
        if v is None or v == "":
            out.append(np.nan)
        else:
            try:
                out.append(float(v))
            except (TypeError, ValueError):
                out.append(np.nan)
    return np.array(out, dtype=np.float64)


def load_excel(path: str, skip: int) -> dict:
    # Load data from our standard YOLO Excel format.
    # Bizim standart YOLO Excel formatimizdan veri yukler.
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Training Metrics"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty sheet in {path}")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]

    def col(name: str) -> np.ndarray:
        if name not in headers:
            raise KeyError(f"Column not found in Excel: {name}")
        idx = headers.index(name)
        return _to_float_array([r[idx] for r in data_rows])

    data = {
        "epoch":    col("Epoch"),
        "t_box":    col("Box Loss (Train)"),
        "v_box":    col("Box Loss (Val)"),
        "t_cls":    col("Class Loss (Train)"),
        "v_cls":    col("Class Loss (Val)"),
        "t_dfl":    col("DFL Loss (Train)"),
        "v_dfl":    col("DFL Loss (Val)"),
        "prec":     col("Precision (Val)"),
        "rec":      col("Recall (Val)"),
        "map50":    col("mAP@0.5 (overall)"),
        "map5095":  col("mAP@0.5:0.95 (overall)"),
    }

    prefix = "mAP@0.5 ("
    class_names = []
    for h in headers:
        if h.startswith(prefix) and h.endswith(")") and h != "mAP@0.5 (overall)":
            class_names.append(h[len(prefix):-1])

    per_class = {}
    for name in class_names:
        per_class[name] = {
            "map50":     col(f"mAP@0.5 ({name})"),
            "map5095":   col(f"mAP@0.5:0.95 ({name})"),
            "precision": col(f"Precision ({name})"),
            "recall":    col(f"Recall ({name})"),
        }
    data["class_names"] = class_names
    data["per_class"]   = per_class

    return _apply_skip(data, skip)


def load_excel_faster_rcnn(path: str, skip: int) -> dict:
    # Load data from Faster R-CNN Excel format.
    # Faster R-CNN Excel formatindan veri yukler.
    #
    # Differences from YOLO format / YOLO formatindan farklari:
    #     - 4 loss components instead of 3 (cls, box_reg, objectness, rpn_box_reg).
    #     - 4 loss component, 3 yerine.
    #     - Validation losses for all 4 components.
    #     - Tum 4 component icin validation loss'lari.
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Training Metrics"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty sheet in {path}")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]

    def col(name: str) -> np.ndarray:
        if name not in headers:
            raise KeyError(f"Column not found in Excel: {name}")
        idx = headers.index(name)
        return _to_float_array([r[idx] for r in data_rows])

    data = {
        "epoch":    col("Epoch"),
        "t_cls":    col("Loss Classifier (Train)"),
        "t_box":    col("Loss Box Reg (Train)"),
        "t_obj":    col("Loss Objectness (Train)"),
        "t_rpn":    col("Loss RPN Box Reg (Train)"),
        "v_cls":    col("Loss Classifier (Val)"),
        "v_box":    col("Loss Box Reg (Val)"),
        "v_obj":    col("Loss Objectness (Val)"),
        "v_rpn":    col("Loss RPN Box Reg (Val)"),
        "prec":     col("Precision (Val)"),
        "rec":      col("Recall (Val)"),
        "map50":    col("mAP@0.5 (overall)"),
        "map5095":  col("mAP@0.5:0.95 (overall)"),
    }

    prefix = "mAP@0.5 ("
    class_names = []
    for h in headers:
        if h.startswith(prefix) and h.endswith(")") and h != "mAP@0.5 (overall)":
            class_names.append(h[len(prefix):-1])

    per_class = {}
    for name in class_names:
        per_class[name] = {
            "map50":     col(f"mAP@0.5 ({name})"),
            "map5095":   col(f"mAP@0.5:0.95 ({name})"),
            "precision": col(f"Precision ({name})"),
            "recall":    col(f"Recall ({name})"),
        }
    data["class_names"] = class_names
    data["per_class"]   = per_class

    return _apply_skip(data, skip)


def load_yolo_csv(path: str, skip: int) -> dict:
    # Load data from Ultralytics results.csv format.
    # Ultralytics results.csv formatindan veri yukler.
    suffix = Path(path).suffix.lower()
    if suffix in (".xlsx", ".xls"):
        import pandas as pd
        df = pd.read_excel(path)
        df.columns = [c.strip() for c in df.columns]
        rows = [{k: str(v) for k, v in row.items()} for row in df.to_dict(orient="records")]
    else:
        with open(path, newline="") as f:
            raw = f.read()
        first_line = raw.split("\n")[0]
        delimiter = "\t" if "\t" in first_line else ","
        reader = csv.DictReader(io.StringIO(raw), delimiter=delimiter)
        reader.fieldnames = [h.strip() for h in reader.fieldnames]
        rows = [{k.strip(): v.strip() for k, v in row.items() if k} for row in reader]

    def col(key: str) -> np.ndarray:
        return _to_float_array([r[key] for r in rows])

    data = {
        "epoch":    col("epoch"),
        "t_box":    col("train/box_loss"),
        "t_cls":    col("train/cls_loss"),
        "t_dfl":    col("train/dfl_loss"),
        "v_box":    col("val/box_loss"),
        "v_cls":    col("val/cls_loss"),
        "v_dfl":    col("val/dfl_loss"),
        "prec":     col("metrics/precision(B)"),
        "rec":      col("metrics/recall(B)"),
        "map50":    col("metrics/mAP50(B)"),
        "map5095":  col("metrics/mAP50-95(B)"),
        "class_names": [],
        "per_class":   {},
    }

    return _apply_skip(data, skip)


def _apply_skip(data: dict, skip: int) -> dict:
    mask = data["epoch"] > skip
    result = {}
    for k, v in data.items():
        if k == "class_names":
            result[k] = v
        elif k == "per_class":
            result[k] = {
                name: {mk: mv[mask] for mk, mv in metrics.items()}
                for name, metrics in v.items()
            }
        else:
            result[k] = v[mask]
    return result


# Plot helpers.
# Cizim yardimcilari.

def _style_ax(ax, ep, title, ylabel="Loss"):
    ax.set_title(title)
    ax.set_xlabel("Epoch")
    ax.set_ylabel(ylabel)
    if len(ep) > 0:
        ax.set_xlim(ep[0], ep[-1])
    ax.legend()


def plot_single_loss(ax, ep, train, val, title,
                     train_color=TRAIN_COLOR, val_color=VAL_COLOR):
    ax.plot(ep, train, color=train_color, linewidth=LW, label="Train Loss")
    ax.plot(ep, val,   color=val_color,   linewidth=LW,
            linestyle="--", label="Validation Loss")
    _style_ax(ax, ep, title)


# YOLO overall plots / YOLO genel grafikleri.

def save_box_loss(d, out, dpi, fmt):
    fig, ax = plt.subplots(figsize=(9, 5))
    plot_single_loss(ax, d["epoch"], d["t_box"], d["v_box"], "Box Loss")
    plt.tight_layout()
    path = out / f"loss_box.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_cls_loss(d, out, dpi, fmt):
    fig, ax = plt.subplots(figsize=(9, 5))
    plot_single_loss(ax, d["epoch"], d["t_cls"], d["v_cls"], "Class Loss",
                     train_color=CLS_COLOR, val_color="#FC8D59")
    plt.tight_layout()
    path = out / f"loss_cls.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_dfl_loss(d, out, dpi, fmt):
    fig, ax = plt.subplots(figsize=(9, 5))
    plot_single_loss(ax, d["epoch"], d["t_dfl"], d["v_dfl"], "DFL Loss",
                     train_color=DFL_COLOR, val_color="#74C476")
    plt.tight_layout()
    path = out / f"loss_dfl.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_three_combined(d, out, dpi, fmt):
    ep = d["epoch"]
    fig, ax = plt.subplots(figsize=(11, 6))

    ax.plot(ep, d["t_box"], color="#2166AC", linewidth=LW, label="Box - Train")
    ax.plot(ep, d["v_box"], color="#2166AC", linewidth=LW,
            linestyle="--", label="Box - Val")

    ax.plot(ep, d["t_cls"], color="#D73027", linewidth=LW, label="Cls - Train")
    ax.plot(ep, d["v_cls"], color="#D73027", linewidth=LW,
            linestyle="--", label="Cls - Val")

    ax.plot(ep, d["t_dfl"], color="#1A9850", linewidth=LW, label="DFL - Train")
    ax.plot(ep, d["v_dfl"], color="#1A9850", linewidth=LW,
            linestyle="--", label="DFL - Val")

    ax.set_title("Box + Class + DFL Loss", fontsize=13, fontweight="bold")
    ax.set_xlabel("Epoch")
    ax.set_ylabel("Loss")
    if len(ep) > 0:
        ax.set_xlim(ep[0], ep[-1])
    ax.legend(ncol=3, loc="upper right")

    plt.tight_layout()
    path = out / f"loss_three_combined.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_combined_dashboard(d, out, dpi, fmt):
    ep = d["epoch"]
    fig = plt.figure(figsize=(18, 10))
    fig.suptitle("Training Metrics / Egitim Metrikleri",
                 fontsize=16, fontweight="bold", y=1.01)
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.42, wspace=0.32)

    for col_idx, (key_t, key_v, title, tc, vc) in enumerate([
        ("t_box", "v_box", "Box Loss",   "#2166AC", "#D73027"),
        ("t_cls", "v_cls", "Class Loss", "#D73027", "#FC8D59"),
        ("t_dfl", "v_dfl", "DFL Loss",   "#1A9850", "#74C476"),
    ]):
        ax = fig.add_subplot(gs[0, col_idx])
        ax.plot(ep, d[key_t], color=tc, linewidth=LW, label="Train Loss")
        ax.plot(ep, d[key_v], color=vc, linewidth=LW,
                linestyle="--", label="Validation Loss")
        _style_ax(ax, ep, title)

    ax4 = fig.add_subplot(gs[1, 0])
    ax4.plot(ep, d["prec"], color="#2166AC", linewidth=LW,
             label="Validation Precision")
    _style_ax(ax4, ep, "Precision", ylabel="Precision")

    ax5 = fig.add_subplot(gs[1, 1])
    ax5.plot(ep, d["rec"], color="#2166AC", linewidth=LW,
             label="Validation Recall")
    _style_ax(ax5, ep, "Recall", ylabel="Recall")

    ax6 = fig.add_subplot(gs[1, 2])
    ax6.plot(ep, d["map50"],   color="#2166AC", linewidth=LW,
             label="mAP@0.5")
    ax6.plot(ep, d["map5095"], color="#D73027", linewidth=LW,
             linestyle="--", label="mAP@0.5:0.95")
    _style_ax(ax6, ep, "mAP", ylabel="mAP")

    path = out / f"loss_combined.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


# Faster R-CNN overall plots / Faster R-CNN genel grafikleri.

def save_frcnn_classifier_loss(d, out, dpi, fmt):
    fig, ax = plt.subplots(figsize=(9, 5))
    plot_single_loss(ax, d["epoch"], d["t_cls"], d["v_cls"],
                     "Classifier Loss",
                     train_color=FRCNN_COLORS["classifier"],
                     val_color="#FC8D59")
    plt.tight_layout()
    path = out / f"loss_classifier.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_frcnn_box_reg_loss(d, out, dpi, fmt):
    fig, ax = plt.subplots(figsize=(9, 5))
    plot_single_loss(ax, d["epoch"], d["t_box"], d["v_box"],
                     "Box Regression Loss",
                     train_color=FRCNN_COLORS["box_reg"],
                     val_color="#92C5DE")
    plt.tight_layout()
    path = out / f"loss_box_reg.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_frcnn_objectness_loss(d, out, dpi, fmt):
    fig, ax = plt.subplots(figsize=(9, 5))
    plot_single_loss(ax, d["epoch"], d["t_obj"], d["v_obj"],
                     "Objectness Loss",
                     train_color=FRCNN_COLORS["objectness"],
                     val_color="#74C476")
    plt.tight_layout()
    path = out / f"loss_objectness.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_frcnn_rpn_box_reg_loss(d, out, dpi, fmt):
    fig, ax = plt.subplots(figsize=(9, 5))
    plot_single_loss(ax, d["epoch"], d["t_rpn"], d["v_rpn"],
                     "RPN Box Regression Loss",
                     train_color=FRCNN_COLORS["rpn_box_reg"],
                     val_color="#C5B0D5")
    plt.tight_layout()
    path = out / f"loss_rpn_box_reg.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_frcnn_four_combined(d, out, dpi, fmt):
    # All 4 Faster R-CNN losses (train and val) on a single plot.
    # Tum 4 Faster R-CNN loss'i (train ve val) tek grafikte.
    ep = d["epoch"]
    fig, ax = plt.subplots(figsize=(12, 6))

    ax.plot(ep, d["t_cls"], color=FRCNN_COLORS["classifier"], linewidth=LW,
            label="Classifier - Train")
    ax.plot(ep, d["v_cls"], color=FRCNN_COLORS["classifier"], linewidth=LW,
            linestyle="--", label="Classifier - Val")

    ax.plot(ep, d["t_box"], color=FRCNN_COLORS["box_reg"], linewidth=LW,
            label="Box Reg - Train")
    ax.plot(ep, d["v_box"], color=FRCNN_COLORS["box_reg"], linewidth=LW,
            linestyle="--", label="Box Reg - Val")

    ax.plot(ep, d["t_obj"], color=FRCNN_COLORS["objectness"], linewidth=LW,
            label="Objectness - Train")
    ax.plot(ep, d["v_obj"], color=FRCNN_COLORS["objectness"], linewidth=LW,
            linestyle="--", label="Objectness - Val")

    ax.plot(ep, d["t_rpn"], color=FRCNN_COLORS["rpn_box_reg"], linewidth=LW,
            label="RPN Box Reg - Train")
    ax.plot(ep, d["v_rpn"], color=FRCNN_COLORS["rpn_box_reg"], linewidth=LW,
            linestyle="--", label="RPN Box Reg - Val")

    ax.set_title("All Faster R-CNN Losses", fontsize=13, fontweight="bold")
    ax.set_xlabel("Epoch")
    ax.set_ylabel("Loss")
    if len(ep) > 0:
        ax.set_xlim(ep[0], ep[-1])
    ax.legend(ncol=4, loc="upper right", fontsize=9)

    plt.tight_layout()
    path = out / f"loss_four_combined.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_frcnn_combined_dashboard(d, out, dpi, fmt):
    # Eight-panel dashboard for Faster R-CNN: 4 losses on top, P/R/mAP on bottom.
    # Faster R-CNN icin sekiz panelli dashboard: ust satir 4 loss, alt satir P/R/mAP.
    ep = d["epoch"]
    fig = plt.figure(figsize=(20, 10))
    fig.suptitle("Faster R-CNN Training Metrics / Faster R-CNN Egitim Metrikleri",
                 fontsize=16, fontweight="bold", y=1.01)
    gs = gridspec.GridSpec(2, 4, figure=fig, hspace=0.42, wspace=0.32)

    panels = [
        ("t_cls", "v_cls", "Classifier Loss",         FRCNN_COLORS["classifier"],  "#FC8D59"),
        ("t_box", "v_box", "Box Regression Loss",     FRCNN_COLORS["box_reg"],     "#92C5DE"),
        ("t_obj", "v_obj", "Objectness Loss",         FRCNN_COLORS["objectness"],  "#74C476"),
        ("t_rpn", "v_rpn", "RPN Box Regression Loss", FRCNN_COLORS["rpn_box_reg"], "#C5B0D5"),
    ]

    for col_idx, (key_t, key_v, title, tc, vc) in enumerate(panels):
        ax = fig.add_subplot(gs[0, col_idx])
        ax.plot(ep, d[key_t], color=tc, linewidth=LW, label="Train Loss")
        ax.plot(ep, d[key_v], color=vc, linewidth=LW,
                linestyle="--", label="Validation Loss")
        _style_ax(ax, ep, title)

    ax_p = fig.add_subplot(gs[1, 0])
    ax_p.plot(ep, d["prec"], color="#2166AC", linewidth=LW,
              label="Validation Precision")
    _style_ax(ax_p, ep, "Precision", ylabel="Precision")

    ax_r = fig.add_subplot(gs[1, 1])
    ax_r.plot(ep, d["rec"], color="#2166AC", linewidth=LW,
              label="Validation Recall")
    _style_ax(ax_r, ep, "Recall", ylabel="Recall")

    ax_m = fig.add_subplot(gs[1, 2])
    ax_m.plot(ep, d["map50"],   color="#2166AC", linewidth=LW,
              label="mAP@0.5")
    ax_m.plot(ep, d["map5095"], color="#D73027", linewidth=LW,
              linestyle="--", label="mAP@0.5:0.95")
    _style_ax(ax_m, ep, "mAP", ylabel="mAP")

    path = out / f"loss_combined.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


# Per-class plots (shared) / Per-class grafikleri (ortak).

def _save_per_class_metric(d, out, dpi, fmt, metric_key, title, ylabel, filename):
    ep = d["epoch"]
    class_names = d["class_names"]
    per_class = d["per_class"]

    if not class_names:
        print(f"  Skipped {filename}: no per-class data found.")
        return

    fig, ax = plt.subplots(figsize=(10, 6))
    for i, name in enumerate(class_names):
        color = CLASS_COLORS[i % len(CLASS_COLORS)]
        ax.plot(ep, per_class[name][metric_key],
                color=color, linewidth=LW, label=name)

    ax.set_title(title)
    ax.set_xlabel("Epoch")
    ax.set_ylabel(ylabel)
    if len(ep) > 0:
        ax.set_xlim(ep[0], ep[-1])
    ax.legend(loc="best")

    plt.tight_layout()
    path = out / f"{filename}.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {path}")


def save_per_class_plots(d, out, dpi, fmt):
    _save_per_class_metric(d, out, dpi, fmt,
        "map50",     "Per-Class mAP@0.5",      "mAP@0.5",      "per_class_map50")
    _save_per_class_metric(d, out, dpi, fmt,
        "map5095",   "Per-Class mAP@0.5:0.95", "mAP@0.5:0.95", "per_class_map5095")
    _save_per_class_metric(d, out, dpi, fmt,
        "precision", "Per-Class Precision",    "Precision",    "per_class_precision")
    _save_per_class_metric(d, out, dpi, fmt,
        "recall",    "Per-Class Recall",       "Recall",       "per_class_recall")


# Main.
# Ana fonksiyon.

def main():
    parser = argparse.ArgumentParser(
        description="Training metric plot generator (YOLO Excel, YOLO CSV, or Faster R-CNN)"
    )
    parser.add_argument("--input", required=True,
                        help="Input file path / Giris dosyasi yolu")
    parser.add_argument("--format", default="excel",
                        choices=["excel", "yolo", "faster_rcnn"],
                        help="Input format / Giris formati")
    parser.add_argument("--out", default="./plots",
                        help="Output directory / Cikti klasoru")
    parser.add_argument("--skip", type=int, default=3,
                        help="Epochs to skip / Atlanacak epoch sayisi")
    parser.add_argument("--per-class", action="store_true", dest="per_class",
                        help="Also produce per-class plots / Per-class grafikleri de uret")
    parser.add_argument("--dpi", type=int, default=150)
    parser.add_argument("--fmt", default="png", choices=["png", "pdf", "svg"])
    args = parser.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    print(f"Reading input: {args.input} (format={args.format})")
    if args.format == "excel":
        d = load_excel(args.input, skip=args.skip)
    elif args.format == "yolo":
        d = load_yolo_csv(args.input, skip=args.skip)
    elif args.format == "faster_rcnn":
        d = load_excel_faster_rcnn(args.input, skip=args.skip)
    else:
        raise ValueError(f"Unknown format: {args.format}")

    print(f"  Total epochs: {len(d['epoch'])} (first {args.skip} skipped)")
    if d.get("class_names"):
        print(f"  Classes: {d['class_names']}")
    print()

    print("Generating plots...")
    if args.format == "faster_rcnn":
        save_frcnn_classifier_loss(d, out, args.dpi, args.fmt)
        save_frcnn_box_reg_loss(d, out, args.dpi, args.fmt)
        save_frcnn_objectness_loss(d, out, args.dpi, args.fmt)
        save_frcnn_rpn_box_reg_loss(d, out, args.dpi, args.fmt)
        save_frcnn_four_combined(d, out, args.dpi, args.fmt)
        save_frcnn_combined_dashboard(d, out, args.dpi, args.fmt)
    else:
        save_box_loss(d, out, args.dpi, args.fmt)
        save_cls_loss(d, out, args.dpi, args.fmt)
        save_dfl_loss(d, out, args.dpi, args.fmt)
        save_three_combined(d, out, args.dpi, args.fmt)
        save_combined_dashboard(d, out, args.dpi, args.fmt)

    if args.per_class:
        if args.format == "yolo":
            print("  Warning: per-class data not available in YOLO CSV format.")
            print("  Uyari: YOLO CSV formatinda per-class veri bulunmuyor.")
        else:
            save_per_class_plots(d, out, args.dpi, args.fmt)

    print()
    print(f"Done. Plots saved to {args.out}")


if __name__ == "__main__":
    main()