# excel_writer.py
#
# Writes benchmark results to a 4-sheet Excel workbook.
# Benchmark sonuclarini 4 sayfali Excel dosyasina yazar.
#
# Sheets / Sayfalar:
#     1. Summary       - top-line metrics, sorted by mAP@0.5
#     1. Summary       - ust seviye metrikler, mAP@0.5'e gore siralanir
#     2. Per-class     - class-level accuracy detail
#     2. Per-class     - sinif seviyesinde dogruluk detayi
#     3. Speed         - latency statistics
#     3. Speed         - gecikme istatistikleri
#     4. Memory & Size - params, FLOPs, GPU mem, disk size, determinism
#     4. Memory & Size - params, FLOPs, GPU bellek, disk boyutu, determinizm

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_THIN       = Side(border_style="thin", color="BBBBBB")
BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
BEST_FILL   = PatternFill("solid", start_color="C6EFCE")
RANK_FONT   = Font(name="Arial", size=10, bold=True)


def _apply_headers(ws, headers: list, widths: list = None) -> None:
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    if widths:
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"


def _write_row(ws, row_num: int, values: list) -> None:
    alt_fill = PatternFill("solid", start_color="EBF3FB") if row_num % 2 == 0 else None
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row_num, column=ci, value=v)
        c.font = DATA_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill


def _fmt(v, digits=4):
    if v is None:
        return "N/A"
    return round(float(v), digits)


def _write_summary(wb, results: list) -> None:
    # Sheet 1: Summary, sorted by mAP@0.5 descending.
    # Sayfa 1: Ozet, mAP@0.5'e gore azalan siralama.
    ws = wb.active
    ws.title = "Summary"

    headers = [
        "Rank", "Model", "Family",
        "mAP@0.5", "mAP@0.5:0.95",
        "Latency (ms)", "FPS",
        "Params (M)", "Disk (MB)",
    ]
    widths = [6, 48, 10, 12, 14, 14, 10, 12, 12]
    _apply_headers(ws, headers, widths)

    sorted_r = sorted(results, key=lambda x: x.get("map50", 0), reverse=True)

    for i, r in enumerate(sorted_r):
        row = i + 2
        values = [
            i + 1,
            r.get("run_name", ""),
            r.get("family", ""),
            _fmt(r.get("map50")),
            _fmt(r.get("map50_95")),
            _fmt(r.get("latency_mean_ms"), 2),
            _fmt(r.get("fps"), 1),
            _fmt(r.get("params_total_M"), 2),
            _fmt(r.get("disk_size_MB"), 2),
        ]
        _write_row(ws, row, values)

        # Highlight top 3.
        # Ilk 3'u vurgula.
        if i < 3:
            for ci in range(1, len(values) + 1):
                ws.cell(row=row, column=ci).fill = BEST_FILL
                ws.cell(row=row, column=ci).font = RANK_FONT


def _write_per_class(wb, results: list) -> None:
    # Sheet 2: Per-class accuracy.
    # Sayfa 2: Sinif bazinda dogruluk.
    ws = wb.create_sheet("Per-class Accuracy")

    # Collect all class names from first result that has per_class.
    # per_class iceren ilk sonuctan tum sinif isimlerini topla.
    class_names = []
    for r in results:
        pc = r.get("per_class", {})
        if pc:
            class_names = list(pc.keys())
            break

    headers = ["Rank", "Model", "Family"]
    for cn in class_names:
        headers.extend([f"{cn} mAP@0.5", f"{cn} mAP@0.5:0.95"])
    widths = [6, 48, 10] + [16] * (len(class_names) * 2)
    _apply_headers(ws, headers, widths)

    sorted_r = sorted(results, key=lambda x: x.get("map50", 0), reverse=True)

    for i, r in enumerate(sorted_r):
        row = i + 2
        values = [i + 1, r.get("run_name", ""), r.get("family", "")]
        pc = r.get("per_class", {})
        for cn in class_names:
            cls_data = pc.get(cn, {})
            values.append(_fmt(cls_data.get("map50")))
            values.append(_fmt(cls_data.get("map50_95")))
        _write_row(ws, row, values)


def _write_speed(wb, results: list) -> None:
    # Sheet 3: Speed details.
    # Sayfa 3: Hiz detaylari.
    ws = wb.create_sheet("Speed Details")

    headers = [
        "Rank", "Model", "Family",
        "Latency Mean (ms)", "Latency Std (ms)",
        "Latency Min (ms)", "Latency Max (ms)",
        "FPS",
        "Warmup Runs", "Measure Runs",
    ]
    widths = [6, 48, 10, 16, 16, 16, 16, 10, 14, 14]
    _apply_headers(ws, headers, widths)

    sorted_r = sorted(results, key=lambda x: x.get("latency_mean_ms", 9999))

    for i, r in enumerate(sorted_r):
        row = i + 2
        values = [
            i + 1,
            r.get("run_name", ""),
            r.get("family", ""),
            _fmt(r.get("latency_mean_ms"), 2),
            _fmt(r.get("latency_std_ms"), 2),
            _fmt(r.get("latency_min_ms"), 2),
            _fmt(r.get("latency_max_ms"), 2),
            _fmt(r.get("fps"), 1),
            20,
            100,
        ]
        _write_row(ws, row, values)


def _write_memory(wb, results: list) -> None:
    # Sheet 4: Memory and size details.
    # Sayfa 4: Bellek ve boyut detaylari.
    ws = wb.create_sheet("Memory & Size")

    headers = [
        "Rank", "Model", "Family",
        "Params Total (M)", "Params Trainable (M)",
        "FLOPs (G)", "GPU Peak (MB)", "Disk Size (MB)",
        "Deterministic", "Max Diff",
    ]
    widths = [6, 48, 10, 16, 18, 12, 14, 14, 14, 12]
    _apply_headers(ws, headers, widths)

    sorted_r = sorted(results, key=lambda x: x.get("params_total_M", 9999))

    for i, r in enumerate(sorted_r):
        row = i + 2
        det = r.get("deterministic", None)
        det_str = "Yes" if det is True else ("No" if det is False else "N/A")
        values = [
            i + 1,
            r.get("run_name", ""),
            r.get("family", ""),
            _fmt(r.get("params_total_M"), 2),
            _fmt(r.get("params_trainable_M"), 2),
            _fmt(r.get("flops_G"), 2) if r.get("flops_G") is not None else "N/A",
            _fmt(r.get("gpu_mem_peak_MB"), 1),
            _fmt(r.get("disk_size_MB"), 2),
            det_str,
            _fmt(r.get("max_diff"), 8) if r.get("max_diff") is not None else "N/A",
        ]
        _write_row(ws, row, values)


def write_benchmark_excel(results: list, output_path: Path) -> None:
    # Write full benchmark results to a 4-sheet Excel workbook.
    # Tam benchmark sonuclarini 4 sayfali Excel dosyasina yazar.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    _write_summary(wb, results)
    _write_per_class(wb, results)
    _write_speed(wb, results)
    _write_memory(wb, results)

    wb.save(output_path)
    print(f"[Benchmark] Excel saved: {output_path}")